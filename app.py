import os
from dotenv import load_dotenv
load_dotenv()
import cloudinary
import cloudinary.uploader
import cloudinary.utils
import cloudinary.utils
import cloudinary
import cloudinary.uploader
import cloudinary
import cloudinary.utils
import socket


import json
import csv
import uuid
import requests
from datetime import timezone
import pytz

from datetime import datetime, timedelta # Ensure timedelta is imported here
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename # Added this import
from slugify import slugify

from flask import Flask, render_template, redirect, url_for, flash, request, session, jsonify, make_response, send_file
from flask import current_app
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import random
import qrcode
import io
import base64
import string 
from markupsafe import Markup
from openpyxl import load_workbook




# SQLAlchemy Imports
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import Column, Integer, String, Text, Boolean, DateTime, Numeric, ForeignKey, func, or_ 
from sqlalchemy.orm import relationship
from sqlalchemy.exc import IntegrityError 

from flask_migrate import Migrate


# Email Sending
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# CSRF protection
from flask_wtf.csrf import CSRFProtect, generate_csrf

# PDF generation
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from reportlab.lib.units import inch
    from reportlab.lib.pagesizes import A4 # Changed to A4
    from reportlab.lib.colors import HexColor
    # NEW IMPORTS FOR FONT HANDLING
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:
    print("ReportLab not installed. PDF generation features will be disabled.")
    SimpleDocTemplate = None
    Paragraph = None
    Spacer = None
    Table = None
    TableStyle = None
    getSampleStyleSheet = None
    ParagraphStyle = None
    TA_RIGHT = None
    TA_CENTER = None
    TA_LEFT = None
    inch = None
    A4 = None
    HexColor = None
    pdfmetrics = None
    TTFont = None


app = Flask(__name__)

# ---- Custom Jinja filter: floatformat (like Django) ----
@app.template_filter('floatformat')
def floatformat_filter(value, precision=2):
    """
    Usage in templates:
        {{ some_number | floatformat(2) }}
    This will show: 123.45
    """
    try:
        if value is None or value == "":
            return ""
        # Convert to float
        num = float(value)
        # Build a format string like "{:.2f}"
        fmt = "{:." + str(int(precision)) + "f}"
        return fmt.format(num)
    except (ValueError, TypeError):
        # If it can't be converted, just return original value safely
        return value


app.jinja_env.filters['slugify'] = slugify
# --- Configuration ---
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY') or os.environ.get('SECRET_KEY') or 'dev-secret-key'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['RECAPTCHA_SECRET_KEY'] = os.environ.get('RECAPTCHA_SECRET_KEY')
app.config['RECAPTCHA_SITE_KEY'] = os.environ.get('RECAPTCHA_SITE_KEY')
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit for uploads
# Get DATABASE from environment, fall back to SQLite for local development
# This will now look for 'DATABASE' as set in your Render environment
uri = os.environ.get('DATABASE_URL')



# Render's PostgreSQL URL might use 'postgres://', but SQLAlchemy prefers 'postgresql://'
if uri and uri.startswith('postgres://'):
    uri = uri.replace('postgres://', 'postgresql://', 1)

# Use the 'uri' variable for SQLALCHEMY_DATABASE_URI
database_url = os.environ.get('DATABASE_URL')
if not database_url:
    raise ValueError("DATABASE_URL environment variable is not set. Cannot connect to the database.")
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url

# Debug print to confirm correct DB URI is being used
print("Database URI in use:", app.config['SQLALCHEMY_DATABASE_URI'])
# In your Flask app setup (e.g., app.py or config.py)

# app.config['SQLALCHEMY_POOL_RECYCLE'] = 300  # Tell app to reconnect every hour
# app.config['SQLALCHEMY_POOL_PRE_PING'] = True # Tell app to 'hello, are you there?' before talking
# Add this line to handle database connection pooling for Render
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_pre_ping": True,
    "pool_recycle": 299, # Set to slightly less than Render's default timeout (300s)
}

# Email Configuration (Brevo SMTP)
#app.config['MAIL_SERVER'] = os.environ.get('SMTP_SERVER')
#app.config['MAIL_PORT'] = int(os.environ.get('SMTP_PORT', 587))
#app.config['MAIL_USE_TLS'] = True
#app.config['MAIL_USERNAME'] = os.environ.get('SMTP_LOGIN')
#app.config['MAIL_PASSWORD'] = os.environ.get('SMTP_PASSWORD')
#app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('SENDER_EMAIL')
app.config['BREVO_API_KEY'] = os.environ.get('BREVO_API_KEY')
app.config['SENDER_EMAIL'] = os.environ.get('SENDER_EMAIL')




import re

def extract_public_id(url, folder=None, resource_type='upload'):
    """Extracts the public ID (without version/extension) from a Cloudinary URL."""
    # Find the segment after /<resource_type>/<type>/v<version>/
    if resource_type == 'image':
        match = re.search(r'/(?:image|raw|video)/upload/(?:v\d+/)?(.*?)(?:\.\w+)?$', url)
    else: # Default for raw/file/download
        match = re.search(r'/(?:image|raw|video)/upload/(?:v\d+/)?(.*?)(?:\.\w+)?$', url)
        
    if match:
        public_id_with_extension = match.group(1)
        # Remove file extension (.pdf, .jpg, etc.)
        public_id = os.path.splitext(public_id_with_extension)[0]
        return public_id
    return None

# Assuming your invoices are uploaded as 'raw' resource type:
INVOICE_RESOURCE_TYPE = "raw"
# In app.py, replace the existing Cloudinary config block (around line 100) with this:

# --- Cloudinary Configuration (Securely loading from environment) ---
# WARNING: DO NOT COMMIT HARDCODED SECRETS TO PUBLIC REPOSITORIES
CLOUDINARY_CLOUD_NAME_ENV = os.environ['CLOUDINARY_CLOUD_NAME']
CLOUDINARY_API_KEY_ENV    = os.environ['CLOUDINARY_API_KEY']
CLOUDINARY_API_SECRET_ENV = os.environ['CLOUDINARY_API_SECRET']

# -------------------------------------------------------------------

cloudinary.config(
    # Use the variables loaded from the environment
    cloud_name = CLOUDINARY_CLOUD_NAME_ENV,
    api_key = CLOUDINARY_API_KEY_ENV,
    api_secret = CLOUDINARY_API_SECRET_ENV, 
    secure = True
)

# Business Details (for invoices, etc.)
app.config['OUR_BUSINESS_NAME'] = "NaiL Mart India"
app.config['OUR_BUSINESS_ADDRESS'] = "Annapurnna Appartment, New Alipore, Kolkata - 700052"
app.config['OUR_GSTIN'] = "29RUPA1234F1Z5" # Example GSTIN
app.config['OUR_PAN'] = "ABCDE1234F" # Example PAN
app.config['DEFAULT_GST_RATE'] = Decimal('18.00') # Default GST rate for products
# app.config['DEFAULT_SHIPPING_CHARGE'] = Decimal('100.00') # This is now deprecated for per-artwork shipping
db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


login_manager = LoginManager(app)
login_manager.login_view = 'user_login'
login_manager.login_message_category = 'info'
csrf = CSRFProtect(app)

# --- Helper Functions ---
import re
@app.template_filter('ist_time')
def ist_time(value, fmt="%d %b %Y, %I:%M %p"):
    if not value:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)

    ist = pytz.timezone("Asia/Kolkata")
    return value.astimezone(ist).strftime(fmt)


def extract_cloudinary_version(url):
    # Regex to find the version number (v followed by digits)
    match = re.search(r'/v(\d+)/', url)
    return int(match.group(1)) if match else None

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def generate_otp(length=6):
    """Generates a random numeric OTP."""
    return ''.join(random.choices(string.digits, k=length))

def get_products_by_category(category_name):
    with open('artworks.json', 'r') as f:
        artworks = json.load(f)
    # Filter artworks matching this category
    return [art for art in artworks if art.get("category") == category_name]

# ... imports
 # Make sure to add this import if it's not present
# --- CRITICAL INVOICE DOWNLOAD DEFINITIONS (FIXED) ---

# This constant must be defined
INVOICE_RESOURCE_TYPE = "raw" 
INVOICE_FOLDER = "invoices" # Assuming your upload uses this folder
# Assuming INVOICE_RESOURCE_TYPE = "raw" is defined globally (it must be)
# Assuming INVOICE_RESOURCE_TYPE = "raw" is defined globally

def generate_signed_invoice_url(order_id):
    """
    Generates a secure, signed URL for the invoice PDF.
    FIX: Changed type from 'authenticated' to 'upload' to match the likely 
    default upload type while keeping 'sign_url=True' for security.
    """
    from datetime import datetime, timedelta
    import re
    
    order = db.session.get(Order, order_id)
    if not order or not order.invoice_file_url:
        raise ValueError(f"Invoice URL not found for Order ID {order_id}")

    # Extract the full path including the version number (vXXXX/folder/filename)
    match = re.search(r'/(?:raw|image|video)/upload/(.+?)(?:\.\w+)?$', order.invoice_file_url)
    
    if not match:
         raise ValueError(f"Could not extract versioned Public ID from URL: {order.invoice_file_url}")
    
    public_id_for_signing = match.group(1) 

    # 1. Generate the signed URL using the full extracted public ID
    url, options = cloudinary.utils.cloudinary_url(
        public_id_for_signing, 
        resource_type="raw", # Must match how the file was uploaded
        type="upload",       # <-- CRITICAL CHANGE: Use 'upload' but rely on signature
        format="pdf",
        sign_url=True,       # <-- This is what provides the security signature
        attachment=f"Invoice_{order.id}.pdf",
        expires_at=int((datetime.utcnow() + timedelta(minutes=10)).timestamp()) 
    )
    
    # Return the complete URL string
    return url 





# NEW: SequenceCounter model for sequential IDs
class SequenceCounter(db.Model):
    id = db.Column(db.String(50), primary_key=True) # e.g., 'order_id_sequence'
    current_value = db.Column(db.BigInteger, nullable=False)

# MODIFIED: Order ID generation function
def generate_order_id():
    counter_name = 'order_id_sequence'
    # Use a transaction for atomic increment to prevent duplicates in concurrent environments
    with db.session.begin_nested(): 
        counter = db.session.query(SequenceCounter).with_for_update().filter_by(id=counter_name).first()
        if not counter:
            # Initialize if it's the first time.
            # Start from 15416760 so the first increment makes it 15416761
            counter = SequenceCounter(id=counter_name, current_value=15416760) 
            db.session.add(counter)
            db.session.flush() # Ensure it's in the session for the update below
        
        counter.current_value += 1
        new_numeric_part = counter.current_value
        # No need to db.session.add(counter) again if it was already fetched/added and modified
        # The flush() above ensures it's tracked.

    # Format the ID as "OD" + padded 8 digits
    return f"OD{new_numeric_part:08d}"

import requests
import json

def send_email(to_email, subject, body_plain=None, html_body=None,
               attachment_path=None, attachment_name=None):
    """
    FINAL VERSION: Send email using Brevo API.
    Used for OTP, signup, order confirmation, etc.
    Attachments are ignored for now to keep it simple.
    """

    api_key = current_app.config.get("BREVO_API_KEY")
    if not api_key:
        current_app.logger.error("BREVO_API_KEY not set; cannot send email.")
        return False

    url = "https://api.brevo.com/v3/smtp/email"

    # Use your configured sender email (SENDER_EMAIL / MAIL_USERNAME)
    sender_email = (
        current_app.config.get("MAIL_USERNAME")
        or current_app.config.get("SENDER_EMAIL")
        or "no-reply@karthikafutures.com"
    )

    payload = {
        "sender": {
            "name": current_app.config.get("OUR_BUSINESS_NAME", "Karthika Futures"),
            "email": sender_email
        },
        "to": [{"email": to_email}],
        "subject": subject,
    }

    # Prefer HTML if provided; else plain text
    if html_body:
        payload["htmlContent"] = html_body
        if body_plain:
            payload["textContent"] = body_plain
    elif body_plain:
        payload["textContent"] = body_plain
    else:
        payload["textContent"] = " "  # avoid empty body issues

    headers = {
        "accept": "application/json",
        "api-key": api_key,
        "content-type": "application/json"
    }

    try:
        resp = requests.post(url, headers=headers, data=json.dumps(payload))
        if resp.status_code in (200, 201):
            print(f"Brevo email sent successfully to {to_email}")
            return True
        else:
            current_app.logger.error(
                f"Brevo email error ({resp.status_code}): {resp.text}"
            )
            return False
    except Exception as e:
        current_app.logger.error(f"Brevo exception while sending to {to_email}: {e}")
        return False


# --- Database Models ---
class User(db.Model, UserMixin):
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    email = Column(String(120), unique=True, nullable=False)
    phone = Column(String(20), unique=True, nullable=True) # Added phone
    password_hash = Column(Text, nullable=False) 
    full_name = Column(String(100), nullable=True) 
    role = Column(String(20), default='customer', nullable=False) # 'customer', 'admin'
    registration_date = Column(DateTime, default=datetime.utcnow)
    email_verified = Column(Boolean, default=False) # Added for OTP verification

    addresses = relationship('Address', backref='user', lazy=True, cascade="all, delete-orphan")
    orders = relationship('Order', backref='customer', lazy=True, cascade="all, delete-orphan")

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def is_admin(self):
        """Checks if the user has the 'admin' role."""
        return self.role == 'admin'

    def __repr__(self):
        return f"User('{self.email}', '{self.role}')"
    def is_first_time_customer(self):
        """Checks if the user has any completed orders in the database."""
        # You need an 'Order' model linked to 'User' by 'user_id'
        # and a way to know if an order is completed (e.g., status is 'Completed')
        completed_orders_count = Order.query.filter_by(user_id=self.id).filter(
            Order.status.in_(['Completed', 'Shipped', 'Delivered']) # Use your actual completed statuses
        ).count()
        return completed_orders_count == 0
class OTP(db.Model):
    id = Column(Integer, primary_key=True)
    user_id = Column(String(36), ForeignKey('user.id'), nullable=False)
    otp_code = Column(String(6), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    expires_at = Column(DateTime, default=lambda: datetime.utcnow() + timedelta(minutes=10)) # OTP valid for 10 minutes

    user = relationship('User', backref='otps')

    def is_valid(self):
        return datetime.utcnow() < self.expires_at

class Category(db.Model):
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = Column(String(100), unique=True, nullable=False)
    description = Column(Text, nullable=True)
    image = Column(String(255), nullable=True) # Path to category image
    display_order = Column(Integer, default=999)# category sorting 
    artworks = relationship('Artwork', backref='category', lazy=True)
    

    def __repr__(self):
        return f"Category('{self.name}')"

class Artwork(db.Model):
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    sku = Column(String(50), unique=True, nullable=False)
    hsn_code = Column(String(20), nullable=True) # NEW LINE
    hsn_description = Column(String(255), nullable=True)
    name = Column(String(200), nullable=False)
    slug = Column(String(255), unique=True, nullable=False)

    description = Column(Text, nullable=True)
    original_price = Column(Numeric(10, 2), nullable=False) # Price before any options or GST
    discount_price = Column(Numeric(10, 2), nullable=True, default=None) 
    display_order = db.Column(db.Integer, default=999) # Add this line
    # Package weight
    package_weight_grams = db.Column(db.Integer, default=0)
    package_length_cm = db.Column(db.Integer, default=0)
    package_width_cm = db.Column(db.Integer, default=0)
    package_height_cm = db.Column(db.Integer, default=0)

    # New GST fields
    cgst_percentage = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    sgst_percentage = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    igst_percentage = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    ugst_percentage = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    cess_percentage = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False) # NEW: CESS percentage
    gst_type = Column(String(20), default='intra_state', nullable=False) # 'intra_state', 'inter_state', 'union_territory'

    stock = Column(Integer, default=0, nullable=False)
    category_id = Column(String(36), ForeignKey('category.id'), nullable=False)
    images = Column(Text, nullable=True) # Stored as JSON string of image paths
    is_featured = Column(Boolean, default=False, nullable=False)
    is_active = Column(Boolean, default=True, nullable=False)
    custom_options = Column(Text, nullable=True) # Stored as JSON string { "Size": {"A4": 0, "A3": 500}, "Frame": {"None": 0, "Wooden": 1000} }
    shipping_charge = Column(Numeric(10, 2), default=Decimal('0.00'), nullable=False) # NEW: Per-artwork shipping charge

    def get_images_list(self):
        try:
            return json.loads(self.images) if self.images else []
        except json.JSONDecodeError:
            return []
    def get_option_image(self, group_name, option_name):
        """
        Return image URL based on selected option.
        Fallback to first image if not mapped.
        """

        images = self.get_images_list()
        if not images:
            return None

        # HARD MAP (minimal, safe)
        option_image_map = {
            "TYPE": {
                "PROFESSIONAL": images[1] if len(images) > 1 else images[0],
                "STANDARD": images[0]
            }
        }

        return option_image_map.get(group_name, {}).get(option_name, images[0])

    def set_images_list(self, images_list):
        self.images = json.dumps(images_list)

    def get_custom_options_dict(self):
        try:
            return json.loads(self.custom_options) if self.custom_options else {}
        except json.JSONDecodeError:
            return {}

    @property
    def selling_price(self):
        """Returns the discounted price if it's set and lower than the original price, otherwise returns the original price."""
        # Use 'original_price' for comparison, as that is your column name
        if self.discount_price is not None and self.discount_price < self.original_price:
            return self.discount_price
        return self.original_price 

    @property
    def selling_price_incl_gst(self):
        """Calculates the selling price including applicable GST based on gst_type."""
        base_price = self.selling_price
        total_gst_rate = Decimal('0.00')
        if self.gst_type == 'intra_state':
            total_gst_rate = self.cgst_percentage + self.sgst_percentage
        elif self.gst_type == 'inter_state':
            total_gst_rate = self.igst_percentage
        elif self.gst_type == 'union_territory':
            total_gst_rate = self.cgst_percentage + self.ugst_percentage
        
        total_gst_rate += self.cess_percentage # Include CESS in total rate for selling price

        return base_price * (1 + total_gst_rate / 100)

    def __repr__(self):
        return f"Artwork('{self.name}', '{self.sku}')"

class Address(db.Model):
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    user_id = Column(String(36), ForeignKey('user.id'), nullable=False)
    label = Column(String(50), nullable=True) # e.g., "Home", "Work", "Admin Office"
    full_name = Column(String(100), nullable=False)
    phone = Column(String(20), nullable=False)
    address_line1 = Column(String(255), nullable=False)
    address_line2 = Column(String(255), nullable=True)
    city = Column(String(100), nullable=False)
    state = Column(String(100), nullable=False)
    pincode = Column(String(10), nullable=False)
    is_default = Column(Boolean, default=False, nullable=False)

    def __repr__(self):
        return f"Address('{self.full_name}', '{self.city}', '{self.pincode}')"

    # NEW: Method to convert Address object to a dictionary for JSON serialization
    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'label': self.label,
            'full_name': self.full_name,
            'phone': self.phone,
            'address_line1': self.address_line1,
            'address_line2': self.address_line2,
            'city': self.city,
            'state': self.state,
            'pincode': self.pincode,
            'is_default': self.is_default
        }

class Order(db.Model):
    # Changed primary key to use the new function generate_order_id
    id = Column(String(10), primary_key=True, default=generate_order_id) # Max 10 chars (OD + 8 digits)
    user_id = Column(String(36), ForeignKey('user.id'), nullable=False)
    order_date = Column(DateTime, default=datetime.utcnow)
    total_amount = Column(Numeric(10, 2), nullable=False)
    customer_gstin = db.Column(db.String(15), nullable=True)  # <-- 15 number GSTIN of Customer for B2B invoice
    gst_number = Column(String(15), nullable=True) 
    status = Column(String(50), default='Pending Payment', nullable=False) # e.g., Pending Payment, Payment Submitted - Awaiting Verification, Payment Verified – Preparing Order, Shipped, Delivered, Cancelled by User, Cancelled by Admin
    payment_status = Column(String(50), default='pending', nullable=False) # e.g., pending, completed, failed
    shipping_address_id = Column(String(36), ForeignKey('address.id'), nullable=True) # Can be null for direct purchase if address not saved
    shipping_charge = Column(Numeric(10, 2), default=Decimal('0.00'), nullable=False)
    courier = Column(String(100), nullable=True)
    tracking_number = Column(String(100), nullable=True)
    remark = Column(Text, nullable=True) # Admin remarks
    cancellation_reason = Column(Text, nullable=True) # New field for cancellation reason
    payment_screenshot = db.Column(db.String(255), nullable=True)
    email_sent_status = db.Column(db.Boolean, default=False, nullable=False)
    invoice_file_url = db.Column(db.String(500), nullable=True)
    invoice_details = Column(Text, nullable=True) 

    # Invoice details stored as JSON string
    invoice_details = Column(Text, nullable=True) # {business_name, gstin, pan, business_address, invoice_number, invoice_date, billing_address, gst_rate_applied, shipping_charge, final_invoice_amount, invoice_status, is_held_by_admin, cgst_amount, sgst_amount, igst_amount, ugst_amount, cess_amount} # NEW: cess_amount

    items = relationship('OrderItem', backref='order', lazy=True, cascade="all, delete-orphan")
    shipping_address = relationship('Address', foreign_keys=[shipping_address_id])

    # Helper to get customer details for display
    @property
    def customer_name(self):
        return self.customer.full_name if self.customer else 'N/A'

    @property
    def customer_email(self):
        return self.customer.email if self.customer else 'N/A'
    
    @property
    def customer_phone(self):
        return self.customer.phone if self.customer else 'N/A'

    def get_shipping_address(self):
        if self.shipping_address:
            return {
                'full_name': self.shipping_address.full_name,
                'phone': self.shipping_address.phone,
                'address_line1': self.shipping_address.address_line1,
                'address_line2': self.shipping_address.address_line2,
                'city': self.shipping_address.city,
                'state': self.shipping_address.state,
                'pincode': self.shipping_address.pincode
            }
        return {}

    def get_invoice_details(self):
        try:
            return json.loads(self.invoice_details) if self.invoice_details else {}
        except json.JSONDecodeError:
            return {}

    def set_invoice_details(self, details_dict):
        self.invoice_details = json.dumps(details_dict)

    def __repr__(self):
        return f"Order('{self.id}', '{self.customer_name}', '{self.total_amount}', '{self.status}')"

class OrderItem(db.Model):
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    order_id = Column(String(10), ForeignKey('order.id'), nullable=False) # Changed to String(10) to match Order ID
    artwork_id = Column(String(36), ForeignKey('artwork.id'), nullable=False)
    quantity = Column(Integer, nullable=False)
    
    # Store price details at the time of order for historical accuracy
    unit_price_before_gst = Column(Numeric(10, 2), nullable=False) # Price of one unit *including* selected options, *before* GST
    
    # Store applied GST percentages at the time of order
    cgst_percentage_applied = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    sgst_percentage_applied = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    igst_percentage_applied = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    ugst_percentage_applied = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False)
    cess_percentage_applied = Column(Numeric(5, 2), default=Decimal('0.00'), nullable=False) # NEW: CESS percentage applied

    selected_options = Column(Text, nullable=True) # e.g., {"Size": "A4", "Frame": "Wooden"}

    artwork = relationship('Artwork')

    @property
    def total_price_before_gst(self):
        return self.unit_price_before_gst * self.quantity

    @property
    def cgst_amount(self):
        return (self.total_price_before_gst * self.cgst_percentage_applied) / 100

    @property
    def sgst_amount(self):
        return (self.total_price_before_gst * self.sgst_percentage_applied) / 100

    @property
    def igst_amount(self):
        return (self.total_price_before_gst * self.igst_percentage_applied) / 100

    @property
    def ugst_amount(self):
        return (self.total_price_before_gst * self.ugst_percentage_applied) / 100

    @property
    def cess_amount(self): # NEW: CESS amount property
        return (self.total_price_before_gst * self.cess_percentage_applied) / 100

    @property
    def total_gst_amount(self):
        return self.cgst_amount + self.sgst_amount + self.igst_amount + self.ugst_amount + self.cess_amount # Include CESS

    @property
    def total_price_incl_gst(self):
        return self.total_price_before_gst + self.total_gst_amount
    
    def get_selected_options_dict(self):
        try:
            return json.loads(self.selected_options) if self.selected_options else {}
        except json.JSONDecodeError:
            return {}

    def __repr__(self):
        return f"OrderItem('{self.artwork.name}', Quantity: {self.quantity}, Total: {self.total_price_incl_gst})"
class StockLog(db.Model):
    """Tracks every stock change for auditing."""
    # NOTE: This uses the Column/String/Integer imports from SQLAlchemy 
    # at the top of app.py, not db.Column.
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    artwork_id = Column(String(36), ForeignKey('artwork.id'), nullable=False)
    
    # The quantity of stock change (-5 for sale, +10 for restock, etc.)
    change_quantity = Column(Integer, nullable=False) 
    
    # Snapshot of the stock after this change occurred
    current_stock = Column(Integer, nullable=False)   
    
    # e.g., 'SALE', 'ADD', 'RETURN', 'ADJUSTMENT'
    change_type = Column(String(50), nullable=False)  
    
    # Link to the Order (SALE, RETURN)
    # The ForeignKey must match the exact type of Order.id (String(10))
    order_id = Column(String(10), ForeignKey('order.id'), nullable=True) 
    
    # User who made the change (Admin/Customer)
    user_id = Column(String(36), ForeignKey('user.id'), nullable=True) 
    
    remarks = Column(Text, nullable=True) 
    timestamp = Column(DateTime, default=datetime.utcnow)

    # Use 'relationship' directly since it's imported at the top
    artwork = relationship('Artwork', backref=db.backref('stock_logs', lazy=True))
    order = relationship('Order', backref=db.backref('stock_logs', lazy=True)) 
    user = relationship('User', backref=db.backref('stock_logs', lazy=True))
class ContactMessage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    subject = db.Column(db.String(200), nullable=True)
    message = db.Column(db.Text, nullable=False)
    submission_date = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f"<ContactMessage '{self.name} - {self.subject}'>"

ADMIN_ALERT_EMAILS = [
    "subhashes@yahoo.com",
    "subhashes@live.com",
    "rupadeydiamond@gmail.com"
]

from flask_mail import Message

def send_new_order_alert(order):
    try:
        subject = f"🛒 New Order Received – {order.id}"

        admin_dashboard_url = "https://nailmartindia.com/admin-dashboard"
        admin_order_url = f"https://nailmartindia.com/admin/order/{order.id}"


        html_content = f"""
        <h3>🛒 New Order Received</h3>

        <p><strong>Order ID:</strong> {order.id}</p>
        <p><strong>Total Amount:</strong> ₹{order.total_amount}</p>
        <p><strong>Status:</strong> {order.status}</p>
        <p><strong>Placed On:</strong> {order.order_date.strftime('%d %b %Y, %I:%M %p')}</p>

        <hr>

        <p>
        👉 <a href="{admin_order_url}"
        style="
            display:inline-block;
            padding:10px 16px;
            background-color:#16a34a;
            color:#ffffff;
            text-decoration:none;
            border-radius:6px;
            font-weight:600;
        ">
        View This Order
        </a>
        </p>

        <p style="margin-top:10px;">
        <a href="{admin_dashboard_url}"
        style="
            display:inline-block;
            padding:8px 14px;
            background-color:#2563eb;
            color:#ffffff;
            text-decoration:none;
            border-radius:6px;
            font-weight:500;
        ">
        Open Admin Dashboard
        </a>
        </p>

        <p style="margin-top:12px;color:#555;">
        You will be asked to login if not already authenticated.
        </p>
        """


        admin_emails = [
            "subhashes@yahoo.com",
            "subhashes@live.com",
            "rupadeydiamond@gmail.com"
        ]

        for email in admin_emails:
            send_email(
                to_email=email,
                subject=subject,
                html_body=html_content
            )

        current_app.logger.info(f"✅ Admin alert email sent for order {order.id}")

    except Exception as e:
        current_app.logger.error(f"❌ Failed to send admin order alert: {e}", exc_info=True)


# Route to handle form submission
@app.route('/contact', methods=['GET', 'POST'])
def contact():
    # This route now serves both the GET and POST requests
    if request.method == 'POST':
        # Get data from the form
        name = request.form.get('name')
        email = request.form.get('email')
        subject = request.form.get('subject')
        message_content = request.form.get('message')

        # Create a new ContactMessage object and save it to the database
        new_message = ContactMessage(
            name=name,
            email=email,
            subject=subject,
            message=message_content
        )
        try:
            db.session.add(new_message)
            db.session.commit()
            flash('Thank you for your message! We will get back to you shortly.', 'success')
        except Exception as e:
            db.session.rollback()
            flash('There was an issue sending your message. Please try again.', 'danger')
            app.logger.error(f"Error submitting contact form: {e}")

        # Redirect the user back to the contact page
        return redirect(url_for('contact'))

    # Your original logic to render the template goes here for GET requests
    # Make sure you have these variables defined in your app's context or configuration.
    our_business_address = app.config.get('OUR_BUSINESS_ADDRESS', 'Annapoorna Appartment, New Alipur, Kolkata - 53')
    our_business_email = app.config.get('MAIL_USERNAME', 'covcorres@gmail.com')
    return render_template('contact.html', our_business_address=our_business_address, our_business_email=our_business_email)

from functools import wraps
from flask import abort

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # In a real app, you would check if the user is an admin.
        # For this example, we'll assume a user with a specific ID is an admin.
        # You should replace this with your actual admin logic.
        if not current_user.is_authenticated or not current_user.is_admin():
            abort(403)  # Forbidden
        return f(*args, **kwargs)
    return decorated_function
# Route for the admin to view messages
# This route is protected so only admins can access it
@app.route('/admin/contact_messages')
@login_required
@admin_required
def admin_contact_messages():
    all_messages = ContactMessage.query.order_by(ContactMessage.submission_date.desc()).all()
    return render_template('admin_contact_messages.html', messages=all_messages)

# --- Flask-Login User Loader ---
@login_manager.user_loader
def load_user(user_id):
    # Fix for LegacyAPIWarning
    return db.session.get(User, user_id)

@app.template_filter('strftime')
def format_datetime(value, format="%Y-%m-%d %H:%M:%S"):
    """Format a datetime object to a string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(format)
    return str(value) # Or raise an error if you only expect datetime objects



# --- Context Processors ---
@app.context_processor
def inject_global_data():
    """Injects data available globally to all templates."""
    cart_count = sum(item['quantity'] for item in session.get('cart', {}).values())
    return {
        'current_year': datetime.now().year,
        'our_business_name': app.config['OUR_BUSINESS_NAME'],
        'our_business_address': app.config['OUR_BUSINESS_ADDRESS'],
        'our_gstin': app.config['OUR_GSTIN'],
        'our_pan': app.config['OUR_PAN'],
        'default_gst_rate': app.config['DEFAULT_GST_RATE'],
        'cart_count': cart_count,
        'now': datetime.utcnow, # For use in templates (e.g., invoice date)
        'timedelta': timedelta # Make timedelta available in Jinja2 context
    }

# --- Custom Jinja2 Filters ---
@app.template_filter('currency')
def currency_filter(value):
    """Formats a number as Indian Rupee currency."""
    try:
        return f"₹{Decimal(value):,.2f}"
    except (InvalidOperation, TypeError):
        return "₹0.00"

@app.template_filter('percent')
def percent_filter(value):
    """Formats a number as a percentage."""
    try:
        return f"{Decimal(value):.2f}%"
    except (InvalidOperation, TypeError):
        return "0.00%"

@app.template_filter('slugify')
def slugify_filter(s):
    """Converts a string to a URL-friendly slug."""
    return slugify(s)

# --- Decorators ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin(): # Changed to call is_admin()
            flash('Admin access required.', 'danger')
            return redirect(url_for('user_login'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/get_cart_count')
def get_cart_count():
    cart_data = session.get('cart', {})
    total_quantity = sum(item['quantity'] for item in cart_data.values())
    return jsonify(success=True, cart_count=total_quantity)

# Helper function to calculate item total including options and GST
def calculate_item_total(artwork, selected_options, quantity):
    """
    Calculates the total price and GST components for a single artwork item,
    considering custom options and quantity.
    Returns a dictionary of detailed calculation results.
    """
    base_price = artwork.selling_price

    # Add option prices if applicable
    options_data = artwork.get_custom_options_dict()
    for option_group, option_value in selected_options.items():
        if option_group in options_data and option_value in options_data[option_group]:
            base_price += Decimal(str(options_data[option_group][option_value]))

    unit_price_before_gst = base_price

    # Use artwork's specific GST percentages
    cgst_percentage = artwork.cgst_percentage
    sgst_percentage = artwork.sgst_percentage
    igst_percentage = artwork.igst_percentage
    ugst_percentage = artwork.ugst_percentage
    cess_percentage = artwork.cess_percentage # NEW: Get CESS percentage

    total_price_before_gst_for_item = unit_price_before_gst * quantity

    cgst_amount = (total_price_before_gst_for_item * cgst_percentage) / 100
    sgst_amount = (total_price_before_gst_for_item * sgst_percentage) / 100
    igst_amount = (total_price_before_gst_for_item * igst_percentage) / 100
    ugst_amount = (total_price_before_gst_for_item * ugst_percentage) / 100
    cess_amount = (total_price_before_gst_for_item * cess_percentage) / 100 # NEW: CESS amount

    total_gst_amount_for_item = cgst_amount + sgst_amount + igst_amount + ugst_amount + cess_amount # Include CESS
    total_price_incl_gst = total_price_before_gst_for_item + total_gst_amount_for_item

    return {
        'unit_price_before_gst': unit_price_before_gst,
        'cgst_percentage_applied': cgst_percentage,
        'sgst_percentage_applied': sgst_percentage,
        'igst_percentage_applied': igst_percentage,
        'ugst_percentage_applied': ugst_percentage,
        'cess_percentage_applied': cess_percentage, # NEW: Pass CESS percentage
        'total_price_incl_gst': total_price_incl_gst,
        'cgst_amount': cgst_amount,
        'sgst_amount': sgst_amount,
        'igst_amount': igst_amount,
        'ugst_amount': ugst_amount,
        'cess_amount': cess_amount, # NEW: Pass CESS amount
        'total_price_before_gst': total_price_before_gst_for_item
    }

# Assuming you have a get_cart_items_details() function somewhere,
from decimal import Decimal, InvalidOperation # Ensure InvalidOperation is imported

def get_cart_items_details():
    """
    Retrieves detailed information for items in the session cart or direct purchase session.
    Prioritizes direct_purchase_cart if it exists.
    Returns (detailed_cart_items, subtotal_before_gst, total_cgst_amount,
              total_sgst_amount, total_igst_amount, total_ugst_amount,
              total_gst_amount, grand_total, total_shipping_charge)
    """
    detailed_cart_items = []
    subtotal_before_gst = Decimal('0.00')
    total_cgst_amount = Decimal('0.00')
    total_sgst_amount = Decimal('0.00')
    total_igst_amount = Decimal('0.00')
    total_ugst_amount = Decimal('0.00')
    total_cess_amount = Decimal('0.00') # NEW: Total CESS amount
    total_gst_amount = Decimal('0.00')
    grand_total = Decimal('0.00')
    total_shipping_charge = Decimal('0.00') # Initialize total shipping charge

    items_source = {}
    if 'direct_purchase_cart' in session and session['direct_purchase_cart']:
        for key, value in session['direct_purchase_cart'].items():
            items_source[key] = value
    else:
        items_source = session.get('cart', {}).copy()

    for item_key, item_data in items_source.items():
        sku = item_data['sku']
        quantity = item_data['quantity']
        selected_options = item_data.get('selected_options', item_data.get('options', {}))

        artwork = Artwork.query.filter_by(sku=sku).first()
        if artwork:
            results_from_calculation = calculate_item_total(artwork, selected_options, quantity)

            unit_price_before_gst = results_from_calculation['unit_price_before_gst']
            cgst_percentage_applied = results_from_calculation['cgst_percentage_applied']
            sgst_percentage_applied = results_from_calculation['sgst_percentage_applied']
            igst_percentage_applied = results_from_calculation['igst_percentage_applied']
            ugst_percentage_applied = results_from_calculation['ugst_percentage_applied']
            cess_percentage_applied = results_from_calculation['cess_percentage_applied'] # NEW
            total_price_incl_gst = results_from_calculation['total_price_incl_gst']
            cgst_amount = results_from_calculation['cgst_amount']
            sgst_amount = results_from_calculation['sgst_amount']
            igst_amount = results_from_calculation['igst_amount']
            ugst_amount = results_from_calculation['ugst_amount']
            cess_amount = results_from_calculation['cess_amount'] # NEW
            total_price_before_gst_for_item = results_from_calculation['total_price_before_gst']

            detailed_cart_items.append({
                'item_key': item_key,
                'artwork': artwork, # Full artwork object
                'quantity': quantity,
                'unit_price_before_gst': unit_price_before_gst,
                'cgst_percentage': cgst_percentage_applied, # CHANGED
                'sgst_percentage': sgst_percentage_applied, # CHANGED
                'igst_percentage': igst_percentage_applied, # CHANGED
                'ugst_percentage': ugst_percentage_applied, # CHANGED
                'cess_percentage': cess_percentage_applied, # NEW
                'total_price_incl_gst': total_price_incl_gst,
                'cgst_amount': cgst_amount,
                'sgst_amount': sgst_amount,
                'igst_amount': igst_amount,
                'ugst_amount': ugst_amount,
                'cess_amount': cess_amount, # NEW
                'selected_options': selected_options,
                'image_url': artwork.get_images_list()[0] if artwork.get_images_list() else 'images/placeholder.png'
            })
            subtotal_before_gst += total_price_before_gst_for_item # CHANGED
            total_cgst_amount += cgst_amount
            total_sgst_amount += sgst_amount
            total_igst_amount += igst_amount
            total_ugst_amount += ugst_amount
            total_cess_amount += cess_amount # NEW
            total_gst_amount += (cgst_amount + sgst_amount + igst_amount + ugst_amount + cess_amount) # Include CESS
            grand_total += total_price_incl_gst
            
            # SHIPPING FIX: charge shipping once per order, using the highest item shipping
            item_shipping_charge = Decimal(str(item_data.get('shipping_charge', artwork.shipping_charge)))

            # Instead of multiplying by quantity, we only keep the MAX shipping charge
            if item_shipping_charge > total_shipping_charge:
                total_shipping_charge = item_shipping_charge

        else:
            flash(f"Artwork with SKU {sku} not found and removed from your cart.", "warning")
            if item_key in session.get('cart', {}):
                del session['cart'][item_key]
            session.modified = True

            # FREE SHIPPING FIX
        FREE_SHIPPING_THRESHOLD = Decimal('999.00')

        if grand_total >= FREE_SHIPPING_THRESHOLD:
            total_shipping_charge = Decimal('0.00')
            
    grand_total += total_shipping_charge

    return (detailed_cart_items, subtotal_before_gst, total_cgst_amount,
            total_sgst_amount, total_igst_amount, total_ugst_amount, total_cess_amount, # NEW: Return total_cess_amount
            total_gst_amount, grand_total, total_shipping_charge)

# NEW: Route to add item to cart (AJAX endpoint)
@app.route('/add-to-cart', methods=['POST'])
@csrf.exempt # Exempt CSRF for AJAX, handled by X-CSRFToken header
def add_to_cart():
    data = request.get_json()
    sku = data.get('sku')
    quantity = int(data.get('quantity', 1))
    selected_options = data.get('selected_options', {}) # Dictionary of selected options

    if not sku or quantity < 1:
        return jsonify(success=False, message='Invalid product or quantity.'), 400

    artwork = Artwork.query.filter_by(sku=sku).first()
    if not artwork:
        return jsonify(success=False, message='Artwork not found.'), 404
    
    if artwork.stock < quantity:
        return jsonify(success=False, message=f'Only {artwork.stock} units of {artwork.name} are available.'), 400

    # Calculate item price including options and GST
    calculated_results = calculate_item_total(artwork, selected_options, quantity)
    
    unit_price_before_gst = calculated_results['unit_price_before_gst']
    cgst_percentage = calculated_results['cgst_percentage_applied']
    sgst_percentage = calculated_results['sgst_percentage_applied']
    igst_percentage = calculated_results['igst_percentage_applied']
    ugst_percentage = calculated_results['ugst_percentage_applied']
    cess_percentage = calculated_results['cess_percentage_applied']

    cart = session.get('cart', {})
    
    # Create a unique key for the cart item based on SKU and selected options
    # This ensures different options for the same SKU are treated as separate cart items
    options_key = json.dumps(selected_options, sort_keys=True)
    item_key = f"{sku}_{options_key}"

    if item_key in cart:
        # If item with same options already in cart, update quantity
        cart[item_key]['quantity'] += quantity
    else:
        # Add new item to cart
        cart[item_key] = {
            'sku': sku,
            'name': artwork.name,
            'imageUrl': artwork.get_images_list()[0] if artwork.get_images_list() else 'images/placeholder.png',
            'quantity': quantity,
            'unitPriceBeforeGst': str(unit_price_before_gst), # Store as string for Decimal
            'cgstPercentage': str(cgst_percentage), # Store as string for Decimal
            'sgstPercentage': str(sgst_percentage),
            'igstPercentage': str(igst_percentage),
            'ugstPercentage': str(ugst_percentage),
            'cessPercentage': str(cess_percentage), # NEW: Store cessPercentage
            'options': selected_options,
            'shippingCharge': str(artwork.shipping_charge)
        }
    
    session['cart'] = cart
    session.modified = True # Important to mark session as modified

    total_quantity_in_cart = sum(item['quantity'] for item in session['cart'].values())
    
    cart_url = url_for('cart') # Get the URL for the cart page
    message_content = f"{quantity} x {artwork.name} added to cart! <a href='{cart_url}' class='alert-link' style='color: inherit; text-decoration: underline; margin-left: 10px;'>Go to Cart</a>"
    flash(Markup(message_content), 'success') # Wrap the message in Markup

    return jsonify(success=True, message='Item added to cart!', cart_count=total_quantity_in_cart)

# NEW: Route to remove item from cart (AJAX endpoint)
@app.route('/remove-from-cart', methods=['POST'])
@csrf.exempt # Exempt CSRF for AJAX, handled by X-CSRFToken header
def remove_from_cart():
    data = request.get_json()
    item_key = data.get('item_key')

    cart = session.get('cart', {})
    if item_key in cart:
        del cart[item_key]
        session['cart'] = cart
        session.modified = True
        total_quantity_in_cart = sum(item['quantity'] for item in session['cart'].values())
        flash('Item removed from cart.', 'info')
        return jsonify(success=True, message='Item removed from cart.', cart_count=total_quantity_in_cart)
    return jsonify(success=False, message='Item not found in cart.'), 404

@app.route('/update-cart-quantity', methods=['POST'])
@csrf.exempt
def update_cart_quantity():
    data = request.get_json()
    item_key = data.get('item_key')
    action = data.get('action')

    cart = session.get('cart', {})

    if item_key not in cart:
        return jsonify(success=False, message="Item not found in cart"), 404

    if action == 'increase':
        cart[item_key]['quantity'] += 1

    elif action == 'decrease':
        cart[item_key]['quantity'] -= 1

        # If quantity becomes 0, remove item
        if cart[item_key]['quantity'] <= 0:
            del cart[item_key]

    session['cart'] = cart
    session.modified = True

    return jsonify(success=True)

@app.route('/cart')
def cart():
    # 1. Calculate all cart totals
    (
        detailed_cart_items,
        subtotal_before_gst,
        total_cgst_amount,
        total_sgst_amount,
        total_igst_amount,
        total_ugst_amount,
        total_cess_amount,
        total_gst_amount,
        grand_total,
        total_shipping_charge
    ) = get_cart_items_details()

    # -----------------------------
    # FREE SHIPPING CONFIG
    # -----------------------------
    FREE_SHIPPING_THRESHOLD = Decimal('999.00')

    # 👉 Amount customer actually pays for products (GST included, shipping excluded)
    payable_amount_for_threshold = grand_total - total_shipping_charge

    # Progress bar %
    # ✅ Calculate order value excluding shipping (but including GST)
    order_value_for_shipping = grand_total - total_shipping_charge

    progress_percentage = min(
        (order_value_for_shipping / FREE_SHIPPING_THRESHOLD) * 100,
        100
    )

    if order_value_for_shipping >= FREE_SHIPPING_THRESHOLD:
        free_shipping_unlocked = True
        amount_needed_for_free_shipping = Decimal('0.00')
    else:
        free_shipping_unlocked = False
        amount_needed_for_free_shipping = (
            FREE_SHIPPING_THRESHOLD - order_value_for_shipping
        )


    # -----------------------------
    # FIRST-TIME DISCOUNT LOGIC
    # -----------------------------
    first_time_discount_amount = Decimal('0.00')

    if current_user.is_authenticated and current_user.is_first_time_customer():
        MIN_PURCHASE = Decimal('1000.00')
        DISCOUNT_RATE = Decimal('0.10')

        if subtotal_before_gst >= MIN_PURCHASE:
            first_time_discount_amount = subtotal_before_gst * DISCOUNT_RATE
            grand_total -= first_time_discount_amount
            flash("🎉 You received a 10% first-time customer discount!", "success")

    # -----------------------------
    # UPSELL PRODUCT (SAFE DEFAULT)
    # -----------------------------
    upsell_product = None
    if not free_shipping_unlocked:
        upsell_product = Artwork.query.filter_by(
            sku='NAIL-GLUE-199'
        ).first()

    # -----------------------------
    # RENDER CART
    # -----------------------------
    return render_template(
        'cart.html',
        cart_items=detailed_cart_items,
        subtotal_before_gst=subtotal_before_gst,
        total_cgst_amount=total_cgst_amount,
        total_sgst_amount=total_sgst_amount,
        total_igst_amount=total_igst_amount,
        total_ugst_amount=total_ugst_amount,
        total_cess_amount=total_cess_amount,
        total_gst_amount=total_gst_amount,
        first_time_discount_amount=first_time_discount_amount,
        grand_total=grand_total,
        shipping_charge=total_shipping_charge,
        free_shipping_unlocked=free_shipping_unlocked,
        progress_percentage=progress_percentage,
        amount_needed_for_free_shipping=amount_needed_for_free_shipping,
        upsell_product=upsell_product
    )


# NEW: Route to create a direct order from "Buy Now"
@app.route('/create_direct_order', methods=['POST'])
@csrf.exempt # Exempt CSRF for AJAX, handled by X-CSRFToken header
def create_direct_order():
    data = request.get_json()
  
    item_to_buy_now = data 

    if not item_to_buy_now:
        return jsonify(success=False, message='No items provided for direct purchase.'), 400

    # Convert numeric values to strings before storing in session to ensure consistency
    # with how cart items are stored (which are initially from form data/DB, not JS floats)
    if 'unitPriceBeforeGst' in item_to_buy_now:
        item_to_buy_now['unitPriceBeforeGst'] = str(item_to_buy_now['unitPriceBeforeGst'])
    if 'cgstPercentage' in item_to_buy_now: # Changed from gstPercentage
        item_to_buy_now['cgstPercentage'] = str(item_to_buy_now['cgstPercentage'])
    if 'sgstPercentage' in item_to_buy_now:
        item_to_buy_now['sgstPercentage'] = str(item_to_buy_now['sgstPercentage'])
    if 'igstPercentage' in item_to_buy_now:
        item_to_buy_now['igstPercentage'] = str(item_to_buy_now['igstPercentage'])
    if 'ugstPercentage' in item_to_buy_now:
        item_to_buy_now['ugstPercentage'] = str(item_to_buy_now['ugstPercentage'])
    if 'cessPercentage' in item_to_buy_now: # NEW
        item_to_buy_now['cessPercentage'] = str(item_to_buy_now['cessPercentage'])

    # NEW: Store shipping_charge in item_to_buy_now if present
    if 'shippingCharge' in item_to_buy_now:
        item_to_buy_now['shippingCharge'] = str(item_to_buy_now['shippingCharge'])

    # Store the direct purchase item in session for later retrieval after login/signup
    # Store as a list of one item to mimic cart structure for purchase_form
    session['direct_purchase_cart'] = {'temp_item_key': item_to_buy_now} 
    session.modified = True

    if current_user.is_authenticated:
        # If logged in, proceed to purchase form directly
        return jsonify(success=True, message='Proceeding to checkout.', redirect_url=url_for('purchase_form'))
    else:
        # If not logged in, redirect to login page, which will then redirect to purchase_form
        # Store the intended destination for after login
        session['redirect_after_auth'] = url_for('purchase_form') 
        session.modified = True
        return jsonify(success=True, message='Please log in or sign up to complete your purchase.', redirect_url=url_for('user_login', next=url_for('purchase_form')))

@app.route('/purchase-form', methods=['GET', 'POST'])
@login_required
def purchase_form():
    # --- 1. Initial Address Setup (Pre-GET/POST) ---
    user_addresses = Address.query.filter_by(user_id=current_user.id).order_by(
        Address.is_default.desc(), Address.id.asc()
    ).all()
    
    prefill_address = None
    if user_addresses:
        for addr in user_addresses:
            if addr.is_default:
                prefill_address = addr
                break
        if not prefill_address:
            prefill_address = user_addresses[0]

    prefill_address_dict = prefill_address.to_dict() if prefill_address else None
    selected_address = prefill_address  # Default selected address

    items_to_process = []
    subtotal_before_gst = Decimal('0.00')
    total_cgst_amount = Decimal('0.00')
    total_sgst_amount = Decimal('0.00')
    total_igst_amount = Decimal('0.00')
    total_ugst_amount = Decimal('0.00')
    total_cess_amount = Decimal('0.00')
    total_gst = Decimal('0.00')
    shipping_charge = Decimal('0.00')
    final_total_amount = Decimal('0.00')
    shipping_address_obj = None

    if request.method == 'POST':
        (
            items_to_process,
            subtotal_before_gst,
            total_cgst_amount,
            total_sgst_amount,
            total_igst_amount,
            total_ugst_amount,
            total_cess_amount,
            total_gst,
            final_total_amount,
            shipping_charge,
        ) = get_cart_items_details()

        if not items_to_process:
            flash("No items to purchase.", "danger")
            return redirect(url_for('cart'))

        # ✅ GST number for both add_new_address and place_order
        gst_number = request.form.get('gst_number', '').strip().upper()
        print(
            "DEBUG purchase_form: gst_number received ->",
            repr(gst_number),
            " action_type ->",
            request.form.get('action_type'),
        )

        # Selected address for re-rendering
        selected_address_id_on_post = request.form.get('selected_address_id') or request.form.get('shipping_address')
        current_selected_address_obj = (
            db.session.get(Address, selected_address_id_on_post)
            if selected_address_id_on_post
            else prefill_address
        )

                # --- GST Validation Error ---
        gst_pattern = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[A-Z0-9]{1}Z[A-Z0-9]{1}$')
        if gst_number and not gst_pattern.match(gst_number):
            flash('Please enter a valid GSTIN (e.g. 19AAABBB1234C1Z4) or leave the field empty.', 'danger')
            return render_template(
                'purchase_form.html',
                items_to_process=items_to_process,
                subtotal_before_gst=subtotal_before_gst,
                total_gst=total_gst,
                shipping_charge=shipping_charge,
                final_total_amount=final_total_amount,
                user_addresses=user_addresses,
                selected_address=current_selected_address_obj,
                prefill_address=prefill_address_dict,
                form_data=request.form.to_dict(),
                total_cgst_amount=total_cgst_amount,
                total_sgst_amount=total_sgst_amount,
                total_igst_amount=total_igst_amount,
                total_ugst_amount=total_ugst_amount,
                total_cess_amount=total_cess_amount,
                has_addresses=bool(user_addresses),
                current_user_data={
                    'full_name': current_user.full_name,
                    'phone': current_user.phone,
                    'email': current_user.email
                }
            )


        action_type = request.form.get('action_type')

        # --- ADD NEW ADDRESS ---
        if action_type == 'add_new_address':
            full_name = request.form.get('full_name')
            phone = request.form.get('phone')
            address_line1 = request.form.get('address_line1')
            address_line2 = request.form.get('address_line2')
            city = request.form.get('city')
            state = request.form.get('state')
            pincode = request.form.get('pincode')
            set_as_default = request.form.get('set_as_default') == 'on'

            if not all([full_name, phone, address_line1, city, state, pincode]):
                flash('Please fill in all required fields for the new address.', 'danger')
                return render_template(
                    'purchase_form.html',
                    items_to_process=items_to_process,
                    subtotal_before_gst=subtotal_before_gst,
                    total_gst=total_gst,
                    shipping_charge=shipping_charge,
                    final_total_amount=final_total_amount,
                    user_addresses=user_addresses,
                    selected_address=current_selected_address_obj,
                    prefill_address=prefill_address_dict,
                    form_data=request.form.to_dict(),
                    total_cgst_amount=total_cgst_amount,
                    total_sgst_amount=total_sgst_amount,
                    total_igst_amount=total_igst_amount,
                    total_ugst_amount=total_ugst_amount,
                    total_cess_amount=total_cess_amount,
                    has_addresses=bool(user_addresses),
                    current_user_data={
                        'full_name': current_user.full_name,
                        'phone': current_user.phone,
                        'email': current_user.email,
                    },
                )

            new_address = Address(
                user_id=current_user.id,
                full_name=full_name,
                phone=phone,
                address_line1=address_line1,
                address_line2=address_line2,
                city=city,
                state=state,
                pincode=pincode,
                is_default=set_as_default,
            )
            if new_address.is_default:
                Address.query.filter_by(user_id=current_user.id, is_default=True).update({"is_default": False})
            db.session.add(new_address)
            db.session.commit()
            session['pre_selected_address_id'] = new_address.id
            flash('New address added successfully! Please select it below before placing the order.', 'info')
            return redirect(url_for('purchase_form'))

        # --- PLACE ORDER ---
        elif action_type == 'place_order':
            selected_address_id = request.form.get('selected_address_id') or request.form.get('shipping_address')

            if not selected_address_id:
                flash('Please select a shipping address.', 'danger')
                return render_template(
                    'purchase_form.html',
                    items_to_process=items_to_process,
                    subtotal_before_gst=subtotal_before_gst,
                    total_gst=total_gst,
                    shipping_charge=shipping_charge,
                    final_total_amount=final_total_amount,
                    user_addresses=user_addresses,
                    selected_address=None,
                    prefill_address=prefill_address_dict,
                    form_data=request.form.to_dict(),
                    total_cgst_amount=total_cgst_amount,
                    total_sgst_amount=total_sgst_amount,
                    total_igst_amount=total_igst_amount,
                    total_ugst_amount=total_ugst_amount,
                    total_cess_amount=total_cess_amount,
                    has_addresses=bool(user_addresses),
                    current_user_data={
                        'full_name': current_user.full_name,
                        'phone': current_user.phone,
                        'email': current_user.email,
                    },
                )

            shipping_address_obj = db.session.get(Address, selected_address_id)

            if not shipping_address_obj or shipping_address_obj.user_id != current_user.id:
                flash("Invalid address selection.", "danger")
                return render_template(
                    'purchase_form.html',
                    items_to_process=items_to_process,
                    subtotal_before_gst=subtotal_before_gst,
                    total_gst=total_gst,
                    shipping_charge=shipping_charge,
                    final_total_amount=final_total_amount,
                    user_addresses=user_addresses,
                    selected_address=None,
                    prefill_address=prefill_address_dict,
                    form_data=request.form.to_dict(),
                    total_cgst_amount=total_cgst_amount,
                    total_sgst_amount=total_sgst_amount,
                    total_igst_amount=total_igst_amount,
                    total_ugst_amount=total_ugst_amount,
                    total_cess_amount=total_cess_amount,
                    has_addresses=bool(user_addresses),
                    current_user_data={
                        'full_name': current_user.full_name,
                        'phone': current_user.phone,
                        'email': current_user.email,
                    },
                )

            try:
                new_order = Order(
                    id=generate_order_id(),
                    user_id=current_user.id,
                    total_amount=final_total_amount,
                    status='Pending Payment',
                    payment_status='pending',
                    shipping_address_id=shipping_address_obj.id,
                    shipping_charge=shipping_charge,
                    customer_gstin=gst_number if gst_number else None,
                    gst_number=gst_number if gst_number else None,
                )
                db.session.add(new_order)
                db.session.flush()

                for item in items_to_process:
                    order_item = OrderItem(
                        order_id=new_order.id,
                        artwork_id=item['artwork'].id,
                        quantity=item['quantity'],
                        unit_price_before_gst=item['unit_price_before_gst'],
                        cgst_percentage_applied=item['cgst_percentage'],
                        sgst_percentage_applied=item['sgst_percentage'],
                        igst_percentage_applied=item['igst_percentage'],
                        ugst_percentage_applied=item['ugst_percentage'],
                        cess_percentage_applied=item['cess_percentage'],
                        selected_options=json.dumps(item['selected_options']),
                    )
                    db.session.add(order_item)

                    artwork = db.session.get(Artwork, item['artwork'].id)
                    if artwork:
                        artwork.stock -= item.get('quantity', 0)
                        if artwork.stock < 0:
                            artwork.stock = 0

                        stock_log = StockLog(
                            artwork_id=artwork.id,
                            change_quantity=-item.get('quantity', 0),
                            current_stock=artwork.stock,
                            change_type='SALE',
                            order_id=new_order.id,
                            user_id=current_user.id,
                            remarks=f"Stock reduced after order {new_order.id}",
                        )
                        db.session.add(stock_log)

                db.session.commit()

                # 🔔 ALERT ADMIN ABOUT NEW ORDER (EMAIL)
                send_new_order_alert(new_order)


                # Clear carts
                session.pop('cart', None)
                session.pop('direct_purchase_cart', None)
                session.modified = True

                # ✅ No Flask-Mail / Message here.
                # Order confirmation email is handled in order_summary using send_email (Brevo).

                flash('Order placed successfully! Please proceed to payment.', 'success')
                return redirect(
                    url_for('payment_initiate', order_id=new_order.id, amount=new_order.total_amount)
                )

            except IntegrityError:
                db.session.rollback()
                flash('An error occurred while creating your order. Please try again.', 'danger')
                return redirect(url_for('purchase_form'))
            except Exception as e:
                db.session.rollback()
                flash(f'An unexpected error occurred: {e}', 'danger')
                return redirect(url_for('purchase_form'))

    # ----- Handle GET -----
    (
        items_to_process,
        subtotal_before_gst,
        total_cgst_amount,
        total_sgst_amount,
        total_igst_amount,
        total_ugst_amount,
        total_cess_amount,
        total_gst,
        final_total_amount,
        shipping_charge,
    ) = get_cart_items_details()

    if not items_to_process:
        flash("No items to purchase.", "danger")
        return redirect(url_for('cart'))

    pre_selected_id = session.pop('pre_selected_address_id', None)
    new_address_added = bool(pre_selected_id)
    if pre_selected_id:
        session_selected_address = db.session.get(Address, pre_selected_id)
        if session_selected_address:
            selected_address = session_selected_address

    form_data = request.form.to_dict() if request.method == 'POST' else {}

    return render_template(
        'purchase_form.html',
        items_to_process=items_to_process,
        subtotal_before_gst=subtotal_before_gst,
        total_gst=total_gst,
        shipping_charge=shipping_charge,
        final_total_amount=final_total_amount,
        user_addresses=user_addresses,
        selected_address=selected_address,
        prefill_address=prefill_address_dict,
        form_data=form_data,
        total_cgst_amount=total_cgst_amount,
        total_sgst_amount=total_sgst_amount,
        total_igst_amount=total_igst_amount,
        total_ugst_amount=total_ugst_amount,
        total_cess_amount=total_cess_amount,
        new_address_added=new_address_added,
        has_addresses=bool(user_addresses),
        current_user_data={
            'full_name': current_user.full_name,
            'phone': current_user.phone,
            'email': current_user.email,
        },
    )

@app.route('/set-default-address/<uuid:address_id>', methods=['POST'])
@login_required
def set_default_address(address_id):
    address_to_set = Address.query.filter_by(id=address_id, user_id=current_user.id).first()

    if not address_to_set:
        flash("Address not found or does not belong to you.", "danger")
        return redirect(url_for('my_addresses'))

    try:
        # Unset the old default address for the user
        current_default = Address.query.filter_by(user_id=current_user.id, is_default=True).first()
        if current_default:
            current_default.is_default = False
        
        # Set the new default address
        address_to_set.is_default = True
        db.session.commit()
        
        flash("Default address has been updated successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"An error occurred while setting the default address: {e}", "danger")

    return redirect(url_for('my_addresses'))
# MODIFIED: Signup route to include OTP verification and next_url capture
# MODIFIED: Signup route to include OTP verification and next_url capture
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    # Store the 'next' URL if provided in the query parameters (e.g., from @login_required)
    next_url = request.args.get('next')
    if next_url:
        session['redirect_after_auth'] = next_url
    # If no 'next' param, try to use referrer, but avoid redirecting back to auth pages
    elif request.referrer and request.referrer != request.url:
        # Check if referrer is not from an authentication page to avoid redirection loops
        if not any(auth_path in request.referrer for auth_path in ['/user-login', '/signup', '/verify_otp', '/forgot-password', '/verify_reset_otp', '/reset_password']):
            session['redirect_after_auth'] = request.referrer
    
    form_data = {}
    
    if request.method == 'POST':
        # THIS IS THE CORRECT LOCATION FOR THE RECAPTCHA CODE
        recaptcha_response = request.form.get('g-recaptcha-response')
        
        # Verify reCAPTCHA
        recaptcha_verify_url = 'https://www.google.com/recaptcha/api/siteverify'
        recaptcha_payload = {
            'secret': app.config['RECAPTCHA_SECRET_KEY'],
            'response': recaptcha_response
        }
        recaptcha_verification = requests.post(recaptcha_verify_url, data=recaptcha_payload)
        recaptcha_result = recaptcha_verification.json()
        
        # Check if reCAPTCHA verification was successful
        if not recaptcha_result.get('success'):
            flash('reCAPTCHA verification failed. Please try again.', 'danger')
            return render_template('signup.html', form_data=request.form)

        # START OF YOUR ORIGINAL CODE
        email = request.form.get('email')
        phone = request.form.get('phone')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        full_name = request.form.get('full_name')

        form_data = {'email': email, 'phone': phone, 'full_name': full_name}

        if not email or not password or not confirm_password:
            flash('Please fill in all required fields.', 'danger')
            return render_template('signup.html', form_data=form_data)

        if password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return render_template('signup.html', form_data=form_data)

        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'danger')
            return render_template('signup.html', form_data=form_data)

        existing_user_email = User.query.filter_by(email=email).first()
        if existing_user_email:
            if existing_user_email.email_verified:
                flash('An account with this email already exists. Please log in.', 'warning')
            else:
                # Resend OTP if email not verified
                otp_code = generate_otp()
                new_otp = OTP(user_id=existing_user_email.id, otp_code=otp_code)
                db.session.add(new_otp)
                db.session.commit()
                send_email(existing_user_email.email, 'Karthika Futures - Verify Your Email', f'Your OTP for email verification is: {otp_code}. It is valid for 10 minutes.')
                session['otp_user_id'] = existing_user_email.id
                flash('An account with this email exists but is not verified. A new OTP has been sent to your email.', 'info')
                return redirect(url_for('verify_otp'))
            return render_template('signup.html', form_data=form_data)
        
        # Check if phone number already exists
        if phone:
            existing_user_phone = User.query.filter_by(phone=phone).first()
            if existing_user_phone:
                flash('An account with this phone number already exists. Please log in or use a different phone number.', 'warning')
                return render_template('signup.html', form_data=form_data)

        # Create user but mark as unverified
        new_user = User(email=email, phone=phone, full_name=full_name, email_verified=False)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()

        # Generate and send OTP
        otp_code = generate_otp()
        otp_entry = OTP(user_id=new_user.id, otp_code=otp_code)
        db.session.add(otp_entry)
        db.session.commit()

        send_email(new_user.email, 'Karthika Futures - Verify Your Email', f'Your OTP for email verification is: {otp_code}. It is valid for 10 minutes.')
        
        session['otp_user_id'] = new_user.id # Store user ID in session for OTP verification
        
        session['prefill_login_phone'] = new_user.phone # Store phone for login prefill

        flash('A One-Time Password (OTP) has been sent to your email. Please verify to complete registration.', 'success')
        return redirect(url_for('verify_otp'))

    return render_template('signup.html', form_data=form_data)

# ... (the rest of your code, including the helper function) ...
# NEW: Helper function to handle post-authentication redirection
def _handle_post_auth_redirect():
    # Prioritize direct purchase flow (from 'Buy Now' button)
    if 'direct_purchase_cart' in session and session['direct_purchase_cart']:
        # Clear the direct purchase cart from session once handled
        session.pop('direct_purchase_cart', None)
        # Ensure 'itemToBuyNow' is cleared if it was set by older logic
        session.pop('itemToBuyNow', None) 
        flash('Login successful! Redirecting to complete your purchase.', 'success')
        return redirect(url_for('purchase_form'))
    
    # Then check for a general 'next' URL (from @login_required or referrer)
    redirect_to_url = session.pop('redirect_after_auth', None)
    if redirect_to_url:
        flash('Login successful! You are now logged in.', 'success')
        return redirect(redirect_to_url)
    
    # Fallback to homepage if no specific URL was stored
    flash('Login successful! You are now logged in.', 'success')
    return redirect(url_for('index'))


@app.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp():
    user_id = session.get('otp_user_id')
    if not user_id:
        flash('No pending verification. Please sign up or try forgot password again.', 'danger')
        return redirect(url_for('signup'))

    # Fix for LegacyAPIWarning
    user = db.session.get(User, user_id)
    if not user:
        flash('User not found for verification.', 'danger')
        return redirect(url_for('signup'))

    if request.method == 'POST':
        otp_entered = request.form.get('otp')
        
        # Find the latest valid OTP for this user
        latest_otp = OTP.query.filter_by(user_id=user.id).order_by(OTP.created_at.desc()).first()

        if latest_otp and latest_otp.otp_code == otp_entered and latest_otp.is_valid():
            user.email_verified = True
            db.session.delete(latest_otp)  # Delete used OTP
            db.session.commit()
            session.pop('otp_user_id', None)

            login_user(user)  # ✅ Automatically log the user in after OTP verified

            # Use the new helper function for redirection
            return _handle_post_auth_redirect()

        else:
            flash('Invalid or expired OTP. Please try again.', 'danger')

    return render_template('verify_otp.html', user_email=user.email)

# NEW: Resend OTP route
@app.route('/resend_otp', methods=['POST'])
def resend_otp():
    user_id = session.get('otp_user_id')
    if not user_id:
        return jsonify(success=False, message='No user session found for OTP resend.'), 400

    # Fix for LegacyAPIWarning
    user = db.session.get(User, user_id)
    if not user:
        return jsonify(success=False, message='User not found.'), 404

    # Generate a new OTP
    otp_code = generate_otp()
    new_otp = OTP(user_id=user.id, otp_code=otp_code)
    db.session.add(new_otp)
    db.session.commit()

    if send_email(user.email, 'Karthika Futures - Your New OTP', f'Your new OTP for verification is: {otp_code}. It is valid for 10 minutes.'):
        return jsonify(success=True, message='New OTP sent to your email.')
    else:
        return jsonify(success=False, message='Failed to send OTP. Please try again later.'), 500


































# MODIFIED: User Login route to check email verification
@app.route('/user-login', methods=['GET', 'POST'])
def user_login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    # Store the 'next' URL if provided in the query parameters (e.g., from @login_required)
    next_url = request.args.get('next')
    if next_url:
        session['redirect_after_auth'] = next_url
    # If no 'next' param, try to use referrer, but avoid redirecting back to auth pages
    elif request.referrer and request.referrer != request.url:
        # Check if referrer is not from an authentication page to avoid redirection loops
        if not any(auth_path in request.referrer for auth_path in ['/user-login', '/signup', '/verify_otp', '/forgot-password', '/verify_reset_otp', '/reset_password']):
            session['redirect_after_auth'] = request.referrer
    
    # Retrieve prefill data from session (will be removed after being read once)
    prefill_phone = session.pop('prefill_login_phone', None)
    prefill_email = session.pop('prefill_login_email', None)

    form_data = {} # Initialize form_data for GET requests or failed POSTs

    if request.method == 'POST':
        # Accept either email or phone for login
        login_identifier = request.form.get('login_identifier')
        password = request.form.get('password')
        form_data = {'login_identifier': login_identifier} # Pass back entered identifier on failure

        # Query for user by either email OR phone number
        user = User.query.filter((User.email == login_identifier) | (User.phone == login_identifier)).first()

        if user and user.check_password(password):
            if not user.email_verified:
                # If email not verified, send OTP and redirect to verify_otp
                otp_code = generate_otp()
                new_otp = OTP(user_id=user.id, otp_code=otp_code)
                db.session.add(new_otp)
                db.session.commit()
                send_email(user.email, 'Karthika Futures - Verify Your Email', f'Your OTP for email verification is: {otp_code}. It is valid for 10 minutes.')
                session['otp_user_id'] = user.id
                flash('Your email is not verified. An OTP has been sent to your email to verify your account.', 'warning')
                return redirect(url_for('verify_otp'))
            
            login_user(user)
            # Use the new helper function for redirection
            return _handle_post_auth_redirect()
        else:
            flash('Invalid login ID or password.', 'danger')

    # Pass prefill data AND form_data to the template for GET requests or failed POST requests
    return render_template('login.html', form_data=form_data, prefill_phone=prefill_phone, prefill_email=prefill_email)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    
    # Clear cartCount from session (if stored)
    session.pop('cartCount', None)
    
    # Clear other session data like 'itemToBuyNow' if needed
    session.pop('itemToBuyNow', None)
    session.pop('redirect_after_login_endpoint', None)
    session.pop('redirect_after_auth', None) # Clear the new redirect key
    session.pop('direct_purchase_cart', None) # Clear direct purchase cart on logout
    
    session.clear()  # Optional: Clears entire session
    
    flash("You’ve been logged out.", "info")
    response = redirect(url_for('index'))
    
    # Optional: Clear cartCount from browser's localStorage via response cookie
    response.set_cookie('cartCount', '', expires=0)  # Helps JS if you sync it with cookies
    return response


# MODIFIED: Forgot Password route to use OTP
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    form_data = {}
    if request.method == 'POST':
        email = request.form.get('email')
        form_data = {'email': email}
        user = User.query.filter_by(email=email).first()

        if user:
            otp_code = generate_otp()
            # Invalidate any previous OTPs for this user for password reset
            OTP.query.filter_by(user_id=user.id).delete()
            db.session.commit()

            new_otp = OTP(user_id=user.id, otp_code=otp_code)
            db.session.add(new_otp)
            db.session.commit()

            send_email(user.email, 'Karthika Futures - Password Reset OTP', f'Your OTP for password reset is: {otp_code}. It is valid for 10 minutes.')
            session['otp_user_id'] = user.id # Store user ID for OTP verification
            flash('A One-Time Password (OTP) has been sent to your email to reset your password.', 'success')
            return redirect(url_for('verify_reset_otp'))
        else:
            flash('No account found with that email address.', 'danger')
    
    return render_template('forgot_password.html', form_data=form_data)

# NEW: Route to verify OTP for password reset
@app.route('/verify_reset_otp', methods=['GET', 'POST'])
def verify_reset_otp():
    user_id = session.get('otp_user_id')
    if not user_id:
        flash('No pending password reset. Please request a new reset.', 'danger')
        return redirect(url_for('forgot_password'))

    # Fix for LegacyAPIWarning
    user = db.session.get(User, user_id)
    if not user:
        flash('User not found for password reset.', 'danger')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        otp_entered = request.form.get('otp')
        
        latest_otp = OTP.query.filter_by(user_id=user.id).order_by(OTP.created_at.desc()).first()

        if latest_otp and latest_otp.otp_code == otp_entered and latest_otp.is_valid():
            user.email_verified = True
            db.session.delete(latest_otp) # Delete used OTP
            db.session.commit()
            # IMPORTANT FIX: Set the user_id in the correct session key for reset_password
            session['reset_user_id'] = user.id
            session.pop('otp_user_id', None) # Clear the temporary OTP user ID
            flash('OTP verified! You can now set your new password.', 'success')
            return redirect(url_for('reset_password'))
        else:
            flash('Invalid or expired OTP. Please try again.', 'danger')
    
    return render_template('verify_reset_otp.html', user_email=user.email)

# NEW: Route to set new password after OTP verification
@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    user_id = session.get('reset_user_id')
    if not user_id:
        flash('Unauthorized access. Please verify OTP first.', 'danger')
        return redirect(url_for('forgot_password'))

    # Fix for LegacyAPIWarning
    user = db.session.get(User, user_id)
    if not user:
        flash('User not found for password reset.', 'danger')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_new_password = request.form.get('confirm_new_password')

        if not new_password or not confirm_new_password:
            flash('Please enter and confirm your new password.', 'danger')
            return render_template('reset_password.html')

        if new_password != confirm_new_password:
            flash('Passwords do not match.', 'danger')
            return render_template('reset_password.html')

        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'danger')
            return render_template('reset_password.html')

        user.set_password(new_password)
        db.session.commit()
        session.pop('reset_user_id', None) # Clear reset user ID
        flash('Your password has been reset successfully! Please log in with your new password.', 'success')
        return redirect(url_for('user_login'))

    return render_template('reset_password.html')


# --- Routes ---
@app.route('/')
def index():
    categories = Category.query.order_by(
        Category.display_order.asc(),
        Category.name.asc()
    ).all()

    all_artworks = Artwork.query.filter(
        Artwork.is_active == True   # ✅ FIX
    ).order_by(Artwork.display_order).all()

    featured_artworks = all_artworks[:10]

    categorized_artworks = defaultdict(list)
    for art in all_artworks:
        if art.category:
            categorized_artworks[art.category.name].append(art)

    return render_template(
        'index.html',
        categories=categories,
        categorized_artworks=categorized_artworks,
        all_artworks=all_artworks,
        featured_artworks=featured_artworks
    )


# NEW: Route for category pages
from datetime import datetime, timedelta
@app.route('/category/<category_slug>')
def category_page(category_slug):
    category = Category.query.filter(
        func.lower(Category.name) == func.lower(category_slug.replace('-', ' '))
    ).first_or_404()

    artworks_in_category = Artwork.query.filter(
        Artwork.category_id == category.id,
        Artwork.is_active == True   # ✅ FIX
    ).order_by(Artwork.display_order).all()

    sale_end_time = datetime.now() + timedelta(hours=6)

    return render_template(
        'category_page.html',
        category=category,
        artworks=artworks_in_category,
        sale_end_time=sale_end_time
    )

@app.route('/all-products')
def all_products():
    search_query = request.args.get('search', '').strip()

    categories = Category.query.order_by(
        Category.display_order.asc(),
        Category.name.asc()
    ).all()

    categorized_artworks = {}

    for category in categories:
        query = Artwork.query.filter(
            Artwork.category_id == category.id,
            Artwork.is_active == True   # ✅ HARD FILTER
        )

        if search_query:
            query = query.filter(
                Artwork.name.ilike(f'%{search_query}%') |
                Artwork.description.ilike(f'%{search_query}%') |
                Artwork.sku.ilike(f'%{search_query}%')
            )

        artworks = query.order_by(Artwork.display_order).all()

        if artworks:
            categorized_artworks[category.name] = artworks

    return render_template(
        'all_products.html',
        categorized_artworks=categorized_artworks,
        search_query=search_query
    )
@app.route('/product/<string:sku>')
def product_detail(sku):
    artwork = Artwork.query.filter_by(
        sku=sku,
        is_active=True
    ).first_or_404()

    artwork_data = {
        'id': artwork.id,
        'sku': artwork.sku,
        'name': artwork.name,
        'description': artwork.description,
        'original_price': float(artwork.original_price),
        'cgst_percentage': float(artwork.cgst_percentage),
        'sgst_percentage': float(artwork.sgst_percentage),
        'igst_percentage': float(artwork.igst_percentage),
        'ugst_percentage': float(artwork.ugst_percentage),
        'cess_percentage': float(artwork.cess_percentage),
        'gst_type': artwork.gst_type,
        'stock': artwork.stock,
        'is_featured': artwork.is_featured,
        'shipping_charge': float(artwork.shipping_charge),
        'image_url': artwork.get_images_list()[0] if artwork.get_images_list() else None,
        'custom_options': artwork.get_custom_options_dict()
    }

    return render_template(
        'product_detail.html',
        artwork=artwork,
        artwork_data=artwork_data
    )


@app.route('/p/<string:slug>')
def product_detail_slug(slug):
    artwork = Artwork.query.filter(
        Artwork.slug == slug,
        Artwork.is_active == True
    ).first_or_404()

    artwork_data = {
        'id': artwork.id,
        'sku': artwork.sku,
        'name': artwork.name,
        'description': artwork.description,
        'original_price': float(artwork.original_price),
        'cgst_percentage': float(artwork.cgst_percentage),
        'sgst_percentage': float(artwork.sgst_percentage),
        'igst_percentage': float(artwork.igst_percentage),
        'ugst_percentage': float(artwork.ugst_percentage),
        'cess_percentage': float(artwork.cess_percentage),
        'gst_type': artwork.gst_type,
        'stock': artwork.stock,
        'is_featured': artwork.is_featured,
        'shipping_charge': float(artwork.shipping_charge),
        'image_url': artwork.get_images_list()[0] if artwork.get_images_list() else None,
        'custom_options': artwork.get_custom_options_dict()
    }

    return render_template(
        'product_detail.html',
        artwork=artwork,
        artwork_data=artwork_data
    )


@app.route('/admin/bulk-upload', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_bulk_upload():

    if request.method == 'POST':
        excel_file = request.files.get('excel_file')

        if not excel_file:
            flash('❌ Please select an Excel file before uploading.', 'danger')
            return redirect(request.url)

        from openpyxl import load_workbook

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
        except Exception:
            flash('❌ Invalid Excel file. Please upload the approved .xlsx format.', 'danger')
            return redirect(request.url)

        headers = [cell.value for cell in ws[1]]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rows.append({k: v if v is not None else '' for k, v in row_dict.items()})


        REQUIRED_COLUMNS = [
            'sku', 'hsn_code', 'hsn_description', 'name', 'description',
            'original_price', 'discount_price', 'stock',
            'shipping_charge',
            'package_weight_grams', 'package_length_cm',
            'package_width_cm', 'package_height_cm',
            'gst_type', 'cgst_percentage', 'sgst_percentage',
            'igst_percentage', 'ugst_percentage', 'cess_percentage',
            'category_id', 'custom_options_json', 'is_featured'
        ]

        missing_columns = [c for c in REQUIRED_COLUMNS if c not in headers]

        if missing_columns:
            flash(
                '❌ Excel format error. Missing columns: ' + ', '.join(missing_columns),
                'danger'
            )
            flash('✔ Fix: Use ONLY the approved bulk-upload Excel template.', 'warning')
            return redirect(request.url)

        preview_rows = rows


        flash(
            f'✔ {len(preview_rows)} rows loaded successfully. Please review before final upload.',
            'success'
        )

        return render_template(
            'admin_bulk_preview.html',
            rows=preview_rows
        )

    return render_template('admin_bulk_upload.html')

@app.route('/admin/bulk-upload-finalize', methods=['POST'])
@login_required
@admin_required
def admin_bulk_finalize():

    try:
        default_image_url = cloudinary.uploader.upload(
            "static/images/default_artwork.jpg",
            folder="artworks/default"
        )['secure_url']
    except Exception:
        flash('❌ Default image upload failed. Check Cloudinary configuration.', 'danger')
        return redirect(url_for('admin_bulk_upload'))

    total = len(request.form.getlist('sku[]'))
    success_count = 0

    for i in range(total):
        sku = request.form.getlist('sku[]')[i].strip()

        # Duplicate SKU check
        if Artwork.query.filter_by(sku=sku).first():
            flash(f'⚠ Duplicate SKU skipped: {sku}', 'warning')
            continue

        try:
            artwork = Artwork(
                sku=sku,
                hsn_code=request.form.getlist('hsn_code[]')[i],
                hsn_description=request.form.getlist('hsn_description[]')[i],
                name=request.form.getlist('name[]')[i],
                description=request.form.getlist('description[]')[i],
                original_price=float(request.form.getlist('original_price[]')[i] or 0),
                discount_price=float(request.form.getlist('discount_price[]')[i] or 0),
                stock=int(request.form.getlist('stock[]')[i] or 0),
                shipping_charge=float(request.form.getlist('shipping_charge[]')[i] or 0),
                package_weight_grams=int(request.form.getlist('package_weight_grams[]')[i] or 0),
                package_length_cm=int(request.form.getlist('package_length_cm[]')[i] or 0),
                package_width_cm=int(request.form.getlist('package_width_cm[]')[i] or 0),
                package_height_cm=int(request.form.getlist('package_height_cm[]')[i] or 0),
                gst_type=request.form.getlist('gst_type[]')[i],
                cgst_percentage=float(request.form.getlist('cgst_percentage[]')[i] or 0),
                sgst_percentage=float(request.form.getlist('sgst_percentage[]')[i] or 0),
                igst_percentage=float(request.form.getlist('igst_percentage[]')[i] or 0),
                ugst_percentage=float(request.form.getlist('ugst_percentage[]')[i] or 0),
                cess_percentage=float(request.form.getlist('cess_percentage[]')[i] or 0),
                category_id=request.form.getlist('category_id[]')[i],
                custom_options=request.form.getlist('custom_options_json[]')[i],
                is_featured=bool(request.form.getlist('is_featured[]')[i]),
                images=json.dumps([default_image_url])
            )
            from slugify import slugify
            artwork.slug = slugify(f"{artwork.name}-{artwork.sku}")

            db.session.add(artwork)
            success_count += 1

        except Exception:
            flash(f'❌ Row skipped due to invalid data (SKU: {sku})', 'danger')
            continue

    db.session.commit()

    flash(
        f'✔ Bulk upload completed successfully. {success_count} artworks added.',
        'success'
    )
    return redirect(url_for('admin_panel'))


@app.route('/order_summary/<order_id>') # This line defines the web address
@login_required # This means only logged-in users can see this page
def order_summary(order_id):
    # --- START OF CHANGES ---
    print(f"DEBUG: Entering order_summary for order_id: {order_id}")
    print(f"DEBUG: Current user ID: {current_user.id}")

    # 1. Find the order in the database
    #    Changed from .first_or_404() to .first() for more direct control and debugging
    #    and added explicit handling if order is None.
    if hasattr(current_user, 'is_admin') and current_user.is_admin:
        order = Order.query.filter_by(id=order_id).first()
    else:
        order = Order.query.filter_by(id=order_id, user_id=current_user.id).first()


    if not order:
        print(f"DEBUG: Order NOT found for ID={order_id} and user ID={current_user.id}.")
        flash('Order not found or you are not authorized to view it.', 'danger')
        return redirect(url_for('user_orders')) # Redirect to a safe page if order isn't found
    
    # If order is found, print its status
    print(f"DEBUG: Order found: ID={order.id}, email_sent_status: {order.email_sent_status}")

    # --- END OF CHANGES ---

    # 2. Check if the email has already been sent for this order
    #    If 'email_sent_status' is False, send the email
    if not order.email_sent_status:
        try:
            # Prepare the email subject
            subject = f"Order Confirmation - Your Order #{order.id} with {current_app.config.get('OUR_BUSINESS_NAME', 'Karthika Futures')}"

            # Get the recipient's email (usually the logged-in user's email)
            recipient_email = current_user.email 

            # 3. Create the HTML content for the email
            #    We use the new template you just created: 'email/order_confirmation.html'
            email_html_body = render_template('email/order_confirmation.html', order=order)

            # 4. Send the email using your updated send_email function
            if send_email(to_email=recipient_email, subject=subject, html_body=email_html_body):
                # If email sent successfully, update the order status in the database
                order.email_sent_status = True
                db.session.commit() # Save the change to the database

              # Hinding email confirmation by #. can acrivate if # removed <<<   flash('Order confirmation email sent successfully!', 'success')
            else:
                
                # Do NOT scare customer. Email failed only because SMTP blocked by host.
                flash(f'Order placed successfully. Please note your Order ID: {order.id}.', 'success')

        except Exception as e:
            # Silent fallback
            flash(f'Order placed successfully. Please note your Order ID: {order.id}.', 'success')
            current_app.logger.error(f"Error sending order confirmation email for Order ID {order.id}: {e}")

    else: # Added for more explicit logging when email is already sent
        print(f"DEBUG: Email already sent for Order ID: {order.id}")


    # 5. Finally, show the order summary page to the user
    return render_template('order_summary.html', order=order)
@app.route('/payment_initiate/<order_id>/<amount>')
@login_required
def payment_initiate(order_id, amount):
    # Fix for LegacyAPIWarning
    order = db.session.get(Order, order_id)
    if not order:
        flash('Order not found.', 'danger')
        return redirect(url_for('user_orders'))

    if order.user_id != current_user.id:
        flash('You are not authorized to make payment for this order.', 'danger')
        return redirect(url_for('user_orders'))

    if order.status != 'Pending Payment':
        flash('This order is not pending payment or has already been processed.', 'warning')
        return redirect(url_for('order_summary', order_id=order.id))

    # Dummy UPI details for demonstration
    upi_id = "smarasada@okaxis" # Replace with your actual UPI ID
    banking_name = "Subhash S" # Replace with your banking name
    

    return render_template('payment_initiate.html', 
                       order=order, # <--- Pass the 'order' object here
                       amount=Decimal(amount), 
                       our_upi_id=upi_id, # Changed variable name to match template
                       our_banking_name=banking_name) # Changed variable name to match template
                       # Removed bank_name as it's not displayed in the template

@app.route('/payment_submit/<order_id>', methods=['POST'])
@login_required
def payment_submit(order_id):
    MAX_SCREENSHOT_SIZE = 1.5 * 1024 * 1024  # 1.5 MB

    order = db.session.get(Order, order_id)
    if not order:
        flash('Order not found.', 'danger')
        return redirect(url_for('user_orders'))

    if order.user_id != current_user.id:
        flash('You are not authorized to submit payment for this order.', 'danger')
        return redirect(url_for('user_orders'))

    if order.status != 'Pending Payment':
        flash('This order is not pending payment or has already been processed.', 'warning')
        return redirect(url_for('order_summary', order_id=order.id))

    payment_screenshot = request.files.get('payment_screenshot')

    if not payment_screenshot:
        flash('Payment screenshot is required to confirm payment.', 'danger')
        return redirect(url_for('payment_initiate', order_id=order.id, amount=order.total_amount))

    if not allowed_file(payment_screenshot.filename):
        flash('Invalid file type for screenshot. Please upload an image.', 'warning')
        return redirect(url_for('payment_initiate', order_id=order.id, amount=order.total_amount))

    # Check file size before saving
    payment_screenshot.seek(0, os.SEEK_END)
    file_size = payment_screenshot.tell()
    payment_screenshot.seek(0)  # reset pointer

    if file_size > MAX_SCREENSHOT_SIZE:
        flash('Screenshot too large. Please upload an image smaller than 1.5 MB.', 'danger')
        return redirect(url_for('payment_initiate', order_id=order.id, amount=order.total_amount))

    # Upload file to Cloudinary and update order
    try:
        cloudinary_result = cloudinary.uploader.upload(payment_screenshot, folder="payment_screenshots")
        if cloudinary_result:
            order.payment_screenshot = cloudinary_result['secure_url']
            order.status = 'Payment Submitted - Awaiting Verification'
            order.payment_status = 'submitted'
            db.session.commit()
            flash('Success! Your order has been placed.', 'success')
            return redirect(url_for('order_summary', order_id=order.id))
        else:
            flash('Failed to upload payment screenshot to Cloudinary. Please try again.', 'danger')
            return redirect(url_for('payment_initiate', order_id=order.id, amount=order.total_amount))
    except Exception as e:
        app.logger.error(f"Cloudinary upload failed for payment screenshot: {e}")
        flash('An error occurred during file upload. Please try again.', 'danger')
        return redirect(url_for('payment_initiate', order_id=order.id, amount=order.total_amount))# NEW: Route for the Thank You page
# CHANGED: <order_id> to <string:order_id>
@app.route('/thank-you/<string:order_id>') 
@login_required # If only logged-in users can view their orders
def thank_you_page(order_id):
    # Fetch the order using filter_by because the ID is a string
    order = Order.query.filter_by(id=order_id).first_or_404()
    
    # Optional: Add security check to ensure user can only view their own order
    if order.user_id != current_user.id:
        flash("You do not have permission to view this order.", "danger")
        return redirect(url_for('user_orders')) # or 'index'
    

    # Insert here to clear cart session
    session['cart'] = {}
    session.modified = True
    
    return render_template('thank_you.html', order=order)


# --- Admin Routes ---
@app.route('/admin-login', methods=['GET', 'POST'])
@csrf.exempt # Exempt CSRF for this route as it's a simple login form
def admin_login():
    if current_user.is_authenticated and current_user.is_admin():
        return redirect(url_for('admin_dashboard'))

    form_data = {}
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        form_data = {'email': email}

        user = User.query.filter_by(email=email, role='admin').first()
        if user and user.check_password(password):
            login_user(user)
            flash('Admin logged in successfully!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid admin credentials.', 'danger')
    return render_template('admin_login.html', form_data=form_data)

@app.route('/admin-dashboard')
@login_required
@admin_required
def admin_dashboard():
    total_users = User.query.count()
    total_artworks = Artwork.query.count()
    total_orders = Order.query.count()
    pending_orders_count = Order.query.filter(Order.status.in_(['Pending Payment', 'Payment Submitted - Awaiting Verification'])).count()

    # Revenue calculation (simplified for dashboard)
    # This sums up total_amount of completed orders
    total_revenue_query = db.session.query(func.sum(Order.total_amount)).filter_by(payment_status='completed').scalar()
    total_revenue = total_revenue_query if total_revenue_query is not None else Decimal('0.00')

    # Orders pending admin review (e.g., payment verification, held invoices)
    orders_pending_review = Order.query.filter(
        (Order.status == 'Payment Submitted - Awaiting Verification') |
        (Order.invoice_details.like('%"is_held_by_admin": true%')) # Check JSON string for held invoices
    ).order_by(Order.order_date.desc()).all()

   # All orders with search and filter (Combined with NEW B2B, Date, and Customer Name filters)
    search_query = request.args.get('search', '')
    filter_status = request.args.get('filter_status', '')

    # NEW: Additional Filters for B2B/B2C, Date Range, and Shipping Name
    filter_b2b = request.args.get('filter_b2b', 'all') 
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    customer_name = request.args.get('customer_name', '').strip()
    
    # Base query for all orders, typically ordered by latest first
    orders_query = Order.query.order_by(Order.order_date.desc())

    # Apply existing filters (Search Query and Status)
    if search_query:
        # Existing logic for searching Order ID, User Name, or Email
        orders_query = orders_query.join(User).filter(
            (Order.id.ilike(f'%{search_query}%')) |
            (User.full_name.ilike(f'%{search_query}%')) |
            (User.email.ilike(f'%{search_query}%'))
        )
    if filter_status:
        orders_query = orders_query.filter_by(status=filter_status)

    # --- START: NEW FILTERING LOGIC (B2B/B2C, Date, Customer Name) ---
    # 1. Apply B2B/B2C filter
    if filter_b2b == 'b2b':
        # B2B: customer_gstin is not null/empty
        orders_query = orders_query.filter(Order.customer_gstin.isnot(None), Order.customer_gstin != '')
    elif filter_b2b == 'b2c':
        # B2C: customer_gstin is null/empty
        # This uses the 'or_' we imported in Step 1!
        orders_query = orders_query.filter(
            or_(Order.customer_gstin.is_(None), Order.customer_gstin == '')
        )

    # 2. Apply Date Range filter
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            orders_query = orders_query.filter(Order.order_date >= start_date)
        if end_date_str:
            # Add one day to include the entire end date in the filter
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
            orders_query = orders_query.filter(Order.order_date < end_date)
    except ValueError:
        flash("Invalid date format provided for filtering.", 'warning')

    # 3. Apply Customer Name filter (using shipping_name)
    if customer_name:
        search_term = f"%{customer_name}%"
        orders_query = orders_query.filter(Order.shipping_name.ilike(search_term))

    # Fetch the final list of orders
    orders = orders_query.all()

    # Pass the applied filters back to the template to persist filter values in the form
    applied_filters = {
        'filter_b2b': filter_b2b,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'customer_name': customer_name
    }
    # --- END: NEW FILTERING LOGIC ---

    # Low stock and out of stock artworks
    low_stock_artworks = Artwork.query.filter(Artwork.stock > 0, Artwork.stock <= 10).all()
    out_of_stock_artworks = Artwork.query.filter_by(stock=0).all()

    # Monthly Revenue for Chart (Last 6 months)
    revenue_labels = []
    revenue_values = []
    for i in range(6, 0, -1):
        month_start = (datetime.utcnow() - timedelta(days=30*i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = (month_start + timedelta(days=30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1) # End of month
        
        monthly_revenue = db.session.query(func.sum(Order.total_amount)).filter(
            Order.order_date >= month_start,
            Order.order_date < month_end,
            Order.payment_status == 'completed'
        ).scalar()
        
        revenue_labels.append(month_start.strftime('%b %Y'))
        revenue_values.append(float(monthly_revenue) if monthly_revenue is not None else 0.0)

    return render_template('admin_dashboard.html',
                           total_users=total_users,
                           total_artworks=total_artworks,
                           total_orders=total_orders,
                           pending_orders=pending_orders_count, # Renamed for clarity
                           total_revenue=total_revenue,
                           orders_pending_review=orders_pending_review,
                           orders=orders,
                           search_query=search_query,
                           filter_status=filter_status,
                           low_stock_artworks=low_stock_artworks,
                           out_of_stock_artworks=out_of_stock_artworks,
                           revenue_labels=revenue_labels,
                           revenue_values=revenue_values,
                           applied_filters=applied_filters)

@app.route('/admin/verify-payment', methods=['POST'])
@login_required
@admin_required
@csrf.exempt # Handled by X-CSRFToken header
def admin_verify_payment():
    order_id = request.form.get('order_id')
    # Fix for LegacyAPIWarning
    order = db.session.get(Order, order_id)
    if order:
        order.status = 'Payment Verified – Preparing Order'
        order.payment_status = 'completed' # Ensure payment status is also completed
        db.session.commit()
        flash(f'Payment for Order ID {order_id} verified successfully!', 'success')
        return jsonify(success=True, message='Payment verified.')
    return jsonify(success=False, message='Order not found.'), 404

@app.route('/admin/order/update_status/<order_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt # Handled by X-CSRFToken header
def admin_update_order_status(order_id): # Added order_id to arguments
    try:
        data = request.get_json()
        # order_id is now correctly passed as a path parameter, no need to get from data
        # Fix for LegacyAPIWarning
        order = db.session.get(Order, order_id)
        if not order:
            return jsonify(success=False, message='Order not found.'), 404

        new_status = data.get('status')
        remark = data.get('remark')
        courier = data.get('courier')
        tracking_number = data.get('tracking_number')
        cancellation_reason = data.get('cancellation_reason') # New: capture cancellation reason

        if new_status:
            order.status = new_status
            order.remark = remark
            order.courier = courier
            order.tracking_number = tracking_number
            order.cancellation_reason = cancellation_reason # Save cancellation reason

            # Update payment status based on order status for consistency
            if 'Cancelled' in new_status:
                order.payment_status = 'cancelled'
            elif new_status == 'Delivered':
                order.payment_status = 'completed' # Assuming payment was completed before delivery
            # Other statuses like 'Shipped', 'Preparing Order' can keep payment_status as 'completed'
            # if it was already marked as such after initial verification.

            db.session.commit()
            flash(f'Order {order_id} status updated to {new_status}.', 'success')
            return jsonify(success=True, message='Order status updated.')
        return jsonify(success=False, message='Invalid status provided.'), 400
    except Exception as e:
        db.session.rollback() # Rollback in case of error
        print(f"Error updating order status for {order_id}: {e}")
        return jsonify(success=False, message='An internal server error occurred.'), 500



@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/delete-user/<user_id>', methods=['POST']) # Corrected from @app.Route
@login_required
@admin_required
def admin_delete_user(user_id):
    # Fix for LegacyAPIWarning
    user = db.session.get(User, user_id)
    if user:
        if user.is_admin(): # Changed to call is_admin()
            flash('Cannot delete an admin user directly.', 'danger')
            return redirect(url_for('admin_users'))
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.email} deleted successfully.', 'success')
    else:
        flash('User not found.', 'danger')
    return redirect(url_for('admin_users'))


















# Route 1: Show stock logs inside admin_dashboard.html tab
@app.route('/admin/stock-logs')
@login_required
@admin_required
def admin_dashboard_stock_logs():
    stock_logs = (
        db.session.query(StockLog, Artwork.name.label('artwork_name'), User.email.label('user_email'))
        .join(Artwork, StockLog.artwork_id == Artwork.id)
        .outerjoin(User, StockLog.user_id == User.id)
        .order_by(StockLog.timestamp.desc())
        .limit(200)
        .all()
    )

    log_entries = []
    for log, artwork_name, user_email in stock_logs:
        log_entries.append({
            'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'artwork_name': artwork_name,
            'change_type': log.change_type,
            'change_quantity': log.change_quantity,
            'current_stock': log.current_stock,
            'order_id': log.order_id if log.order_id else 'N/A',
            'user_email': user_email if user_email else 'System/Admin',
            'remarks': log.remarks or ''
        })

    return render_template(
        'admin_dashboard.html',
        active_tab='stock-history',
        stock_logs=log_entries,
        applied_filters={'filter_b2b': 'all', 'start_date': '', 'end_date': '', 'customer_name': ''},
        orders=[],
        total_users=0,
        total_artworks=0,
        total_orders=0,
        pending_orders=0,
        revenue_labels=[],
        revenue_values=[],
        low_stock_artworks=[],
        out_of_stock_artworks=[],
        orders_pending_review=[]
    )


# Route 2: Separate stock logs page (optional)
@app.route('/admin/stock_logs_page')
@login_required
@admin_required
def admin_stock_logs_page():
    try:
        logs = StockLog.query.order_by(StockLog.timestamp.desc()).all()
        return render_template('admin_stock_logs.html', logs=logs)
    except Exception as e:
        flash(f"Error loading stock logs: {e}", "danger")
        return redirect(url_for('admin_panel'))



@app.route('/admin/order/<order_id>')
@login_required
@admin_required
def admin_order_details(order_id):
    order = Order.query.filter_by(id=order_id).first_or_404()
    return render_template('admin_order_details.html', order=order)

# ... (Keep all existing code above this line)

# --- Administration Routes ---

# Assumes you have an Order model defined, and your existing login_required and admin_required decorators
# ... (Keep all existing code above this line)

# --- Administration Routes ---

# Assumes you have an Order model defined, and your existing login_required and admin_required decorators
@app.route('/admin/orders', methods=['GET'])
@login_required
@admin_required
def admin_orders_view():
    """
    Handles the admin view for listing and filtering orders.
    It processes GET parameters for filtering (e.g., business type, status, dates).
    """
    # Import traceback here if you want detailed error logging
    import traceback
    
    # Check if necessary models and objects are defined.
    # If this causes a NameError, the app will still crash before the try/except block.
    # The error is likely occurring within the try block during query execution.
    
    try:
        # 1. Start with the base query for all orders
        # --- POTENTIAL ERROR SOURCE: Check if Order model and db are correctly imported/accessible ---
        query = Order.query.order_by(Order.order_date.desc())
        
        # 2. Extract and apply filters from the request.form/request.args (since it's a GET form)
        
        # Example filter 1: Business Type (B2B or B2C)
        business_type = request.args.get('business_type', 'all')
        if business_type and business_type != 'all':
            # Assuming you have a 'is_b2b' field on the User or Order model
            # --- POTENTIAL ERROR SOURCE: Check User model and its relationship to Order ---
            if business_type == 'b2b':
                query = query.join(User).filter(User.is_b2b == True)
            elif business_type == 'b2c':
                query = query.join(User).filter(User.is_b2b == False)

        # Example filter 2: Order Status
        status = request.args.get('status', 'all')
        if status and status != 'all':
            # --- POTENTIAL ERROR SOURCE: Check if Order.status is correct column name ---
            query = query.filter(Order.status == status)

        # Example filter 3: Date Range (You'll need more logic for 'start_date' and 'end_date')
        # ...

        # 3. Pagination (Highly recommended for large lists of orders)
        page = request.args.get('page', 1, type=int)
        per_page = 25 # Define how many items per page
        
        # --- POTENTIAL ERROR SOURCE: Query execution happens here ---
        orders = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # 4. Render the template
        return render_template('admin_orders.html', 
                               orders=orders.items, 
                               pagination=orders,
                               current_business_type=business_type,
                               current_status=status)

    except Exception as e:
        # Log the full traceback for detailed error analysis
        current_app.logger.error(f"Error in admin_orders_view: {e}")
        current_app.logger.error(f"FULL TRACEBACK: \n{traceback.format_exc()}") # Enhanced Logging
        
        flash('An error occurred while fetching orders. See console logs for details.', 'danger')
        # Fallback to the main dashboard if there's an error
        return redirect(url_for('admin_dashboard'))

# ... (Keep all existing code below this line)

@app.route('/admin/categories')
@login_required
@admin_required
def admin_categories():
    categories = Category.query.order_by(Category.display_order).all()
    return render_template('admin_categories.html', categories=categories)

@app.route('/admin/add-category', methods=['POST'])
@login_required
@admin_required
def admin_add_category():
    name = request.form.get('category_name')
    display_order = int(request.form.get('display_order', 999))  # ✅ ADD

    if name:
        existing_category = Category.query.filter_by(name=name).first()
        if existing_category:
            flash(f'Category "{name}" already exists.', 'warning')
        else:
            new_category = Category(
                name=name,
                display_order=display_order   # ✅ ADD
            )
            db.session.add(new_category)
            db.session.commit()
            flash(f'Category "{name}" added successfully!', 'success')

    else:
        flash('Category name cannot be empty.', 'danger')
    return redirect(url_for('admin_categories'))

@app.route('/admin/edit-category/<category_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_category(category_id):
    category = Category.query.get_or_404(category_id)
    if request.method == 'POST':
        new_name = request.form.get('name')
        description = request.form.get('description')
        display_order = int(request.form.get('display_order', 999))
        
        if new_name and new_name != category.name:
            existing_category = Category.query.filter_by(name=new_name).first()
            if existing_category and existing_category.id != category.id:
                flash(f'Category name "{new_name}" already exists.', 'danger')
                return render_template('admin_edit_category.html', category=category)
            category.name = new_name
        
        category.display_order = display_order
 

                # Handle image upload (Cloudinary)
        if 'image' in request.files:
            file = request.files['image']

            # Only act if a file was actually chosen
            if file and file.filename:
                if not allowed_file(file.filename):
                    flash('Invalid image file type.', 'danger')
                    return render_template('admin_edit_category.html', category=category)

                try:
                    upload_result = cloudinary.uploader.upload(
                        file,
                        folder="categories"
                    )
                    # Save the secure Cloudinary URL directly in DB
                    category.image = upload_result.get("secure_url")
                except Exception as e:
                    current_app.logger.error(f"Error uploading category image: {e}")
                    flash('Error uploading image. Please try again.', 'danger')
                    return render_template('admin_edit_category.html', category=category)


        db.session.commit()
        flash('Category updated successfully!', 'success')
        return redirect(url_for('admin_categories'))
    return render_template('admin_edit_category.html', category=category)

@app.route('/admin/update-category-order', methods=['POST'])
@login_required
@admin_required
def update_category_order():
    data = request.get_json()

    for item in data:
        category = Category.query.get(item['id'])
        if category:
            category.display_order = item['display_order']

    db.session.commit()
    return jsonify({'success': True})


@app.route('/delete-selected-orders', methods=['POST'])
@login_required

def delete_selected_orders():
    selected_ids = request.form.getlist('selected_orders[]')  # Use the [] version here too
    if not selected_ids:
        flash("No orders selected for deletion.", "warning")
        return redirect(url_for('admin_orders_view'))

    for order_id in selected_ids:
        order = Order.query.filter_by(id=order_id).first()
        if order:
            db.session.delete(order)
    db.session.commit()
    flash(f"{len(selected_ids)} order(s) deleted successfully.", "success")
    return redirect(url_for('admin_orders_view'))


@app.route('/admin/delete-category/<category_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_category(category_id):
    # Fix for LegacyAPIWarning
    category = db.session.get(Category, category_id)
    if category:
        # Check if there are artworks associated with this category
        if category.artworks:
            flash(f'Cannot delete category "{category.name}" because it has associated artworks. Please reassign or delete artworks first.', 'danger')
            return redirect(url_for('admin_categories'))
        
        # Delete category image if exists
        if category.image and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(category.image))):
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(category.image)))

        db.session.delete(category)
        db.session.commit()
        flash(f'Category "{category.name}" deleted successfully.', 'success')
    else:
        flash('Category not found.', 'danger')
    return redirect(url_for('admin_categories'))

@app.route('/admin/artworks')
@login_required
@admin_required
def admin_artworks():
    artworks = Artwork.query.order_by(
        Artwork.display_order,
        Artwork.name
    ).all()

    return render_template('admin_artworks.html', artworks=artworks)

    
@app.route('/admin/toggle-artwork/<string:artwork_id>', methods=['POST'])
@login_required
@admin_required
def toggle_artwork_visibility(artwork_id):
    artwork = Artwork.query.get_or_404(artwork_id)
    artwork.is_active = not artwork.is_active
    db.session.commit()
    flash('Artwork visibility updated.', 'success')
    return redirect(url_for('admin_artworks'))


@app.route('/admin/add-artwork', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_artwork():
    categories = Category.query.all()
    form_data = {}  # Initialize form_data for GET requests

    if request.method == 'POST':
        sku = request.form.get('sku')
        name = request.form.get('name')
        description = request.form.get('description')

        # Use .get() with a default value to prevent NoneType errors
        original_price = request.form.get('original_price', '0.00')

        # New GST fields - provide default '0.00' if not present
        cgst_percentage = request.form.get('cgst_percentage', '0.00')
        sgst_percentage = request.form.get('sgst_percentage', '0.00')
        igst_percentage = request.form.get('igst_percentage', '0.00')
        ugst_percentage = request.form.get('ugst_percentage', '0.00')
        cess_percentage = request.form.get('cess_percentage', '0.00')  # NEW
        gst_type = request.form.get('gst_type')

        stock = request.form.get('stock', '0')  # Default to '0' for stock
        category_id = request.form.get('category_id')
        is_featured = 'is_featured' in request.form  # Checkbox
        

        custom_options_json = request.form.get('custom_options_json') # JSON string from JS

        shipping_charge = request.form.get('shipping_charge', '0.00')  # NEW: Default to '0.00'
        shipping_slab_size = int(request.form.get('shipping_slab_size', 3))

        # Populate form_data for re-rendering on error
        form_data = {
            'sku': sku,
            'name': name,
            'description': description,
            'original_price': original_price,
            'cgst_percentage': cgst_percentage,
            'sgst_percentage': sgst_percentage,
            'igst_percentage': igst_percentage,
            'ugst_percentage': ugst_percentage,
            'cess_percentage': cess_percentage,  # NEW
            'gst_type': gst_type,
            'stock': stock,
            'category_id': category_id,
            'is_featured': is_featured,
            'shipping_charge': shipping_charge
        }

        if custom_options_json:
            try:
                form_data['custom_option_groups'] = json.loads(custom_options_json)
            except json.JSONDecodeError:
                form_data['custom_option_groups'] = {}  # Invalid JSON, treat as empty

        if not all([sku, name, category_id, gst_type]):  # Removed original_price and stock from this check
            flash('Please fill in all required fields (SKU, Name, Category, GST Type).', 'danger')
            return render_template('admin_add_artwork.html', categories=categories, form_data=form_data)

        try:
            original_price = Decimal(original_price)
            cgst_percentage = Decimal(cgst_percentage)
            sgst_percentage = Decimal(sgst_percentage)
            igst_percentage = Decimal(igst_percentage)
            ugst_percentage = Decimal(ugst_percentage)
            cess_percentage = Decimal(cess_percentage)  # NEW
            stock = int(stock)
            shipping_charge = Decimal(shipping_charge)  # NEW: Convert to Decimal
        except (ValueError, InvalidOperation):
            flash('Invalid numeric value for price, GST percentage, stock quantity, or shipping charge.', 'danger')
            return render_template('admin_add_artwork.html', categories=categories, form_data=form_data)

        existing_artwork = Artwork.query.filter_by(sku=sku).first()
        if existing_artwork:
            flash(f'Artwork with SKU "{sku}" already exists.', 'danger')
            return render_template('admin_add_artwork.html', categories=categories, form_data=form_data)

        image_paths = []
        if 'images' in request.files:
            for file in request.files.getlist('images'):
                if file and allowed_file(file.filename):
                    # Upload to Cloudinary and get the public URL
                    upload_result = cloudinary.uploader.upload(file)
                    image_paths.append(upload_result['secure_url'])
                elif file.filename != '':  # If file is present but not allowed
                    flash(f'Invalid file type for image: {file.filename}.', 'danger')
                    return render_template('admin_add_artwork.html', categories=categories, form_data=form_data)

        # --- Create Artwork object ---
        new_artwork = Artwork(
            sku=sku,
            name=name,
            description=description,
            original_price=original_price,
            hsn_code=request.form.get('hsn_code'),          # NEW
            hsn_description=request.form.get('hsn_description'),  # NEW
            cgst_percentage=cgst_percentage,
            sgst_percentage=sgst_percentage,
            igst_percentage=igst_percentage,
            ugst_percentage=ugst_percentage,
            cess_percentage=cess_percentage,                # NEW
            gst_type=gst_type,
            stock=stock,
            category_id=category_id,
            is_featured=is_featured,
            is_active=True,
            shipping_charge=shipping_charge                 # NEW: Assign shipping charge
        )
        from slugify import slugify
        new_artwork.slug = slugify(name)

        new_artwork.set_images_list(image_paths)  # Store image paths as JSON

        # --- NEW: Package weight & dimensions (admin only) ---
        new_artwork.package_weight_grams = int(request.form.get('package_weight_grams') or 0)
        new_artwork.package_length_cm = int(request.form.get('package_length_cm') or 0)
        new_artwork.package_width_cm = int(request.form.get('package_width_cm') or 0)
        new_artwork.package_height_cm = int(request.form.get('package_height_cm') or 0)

        # --- Store custom options as JSON ---
        if custom_options_json:
            try:
                # Validate if it's valid JSON before saving
                json.loads(custom_options_json)
                new_artwork.custom_options = custom_options_json
            except json.JSONDecodeError:
                flash('Invalid format for custom options JSON.', 'danger')
                return render_template('admin_add_artwork.html', categories=categories, form_data=form_data)

        db.session.add(new_artwork)
        db.session.commit()
        flash('Artwork added successfully!', 'success')
        return redirect(url_for('admin_artworks'))

    return render_template('admin_add_artwork.html', categories=categories, form_data=form_data)  # Always pass form_data

@app.route('/admin/edit-artwork/<artwork_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_artwork(artwork_id):
    artwork = Artwork.query.get_or_404(artwork_id)
    categories = Category.query.order_by(Category.display_order).all()


    if request.method == 'POST':
        # --- Retrieve form data ---
        name = request.form.get('name')
        category_id = request.form.get('category_id')
        is_featured = 'is_featured' in request.form
        
        original_price = request.form.get('original_price', '0.00')
        cgst_percentage = request.form.get('cgst_percentage', '0.00')
        sgst_percentage = request.form.get('sgst_percentage', '0.00')
        igst_percentage = request.form.get('igst_percentage', '0.00')
        ugst_percentage = request.form.get('ugst_percentage', '0.00')
        cess_percentage = request.form.get('cess_percentage', '0.00')
        stock = request.form.get('stock', '0')
        description = request.form.get('description')
        display_order = int(request.form.get('display_order', 999))

        shipping_charge = request.form.get('shipping_charge', '0.00')
        gst_type = request.form.get('gst_type')

        # --- NEW: Package details (admin-only) ---
        artwork.package_weight_grams = int(request.form.get('package_weight_grams') or 0)
        artwork.package_length_cm = int(request.form.get('package_length_cm') or 0)
        artwork.package_width_cm = int(request.form.get('package_width_cm') or 0)
        artwork.package_height_cm = int(request.form.get('package_height_cm') or 0)

        images_to_keep = request.form.getlist('images_to_keep')
        custom_options_json_str = request.form.get('custom_options_json')

        # --- Save old stock for logging ---
        old_stock = artwork.stock
        try:
            new_stock = int(stock)
        except ValueError:
            flash('Invalid stock quantity.', 'danger')
            form_data = request.form.to_dict()
            form_data['custom_option_groups'] = json.loads(custom_options_json_str) if custom_options_json_str else {}
            return render_template('admin_edit_artwork.html', artwork=artwork, categories=categories, form_data=form_data)

        # --- Update artwork fields ---
        artwork.name = name
        artwork.slug = slugify(name)

        artwork.category_id = category_id
        artwork.is_featured = is_featured
        
        artwork.description = description

        artwork.gst_type = gst_type
        artwork.hsn_code = request.form.get('hsn_code')
        artwork.hsn_description = request.form.get('hsn_description')

        try:
            artwork.original_price = Decimal(original_price)
            artwork.cgst_percentage = Decimal(cgst_percentage)
            artwork.sgst_percentage = Decimal(sgst_percentage)
            artwork.igst_percentage = Decimal(igst_percentage)
            artwork.ugst_percentage = Decimal(ugst_percentage)
            artwork.cess_percentage = Decimal(cess_percentage)
            artwork.stock = new_stock
            artwork.shipping_charge = Decimal(shipping_charge)
        except (ValueError, InvalidOperation):
            flash('Invalid numeric value for price, GST, stock, or shipping charge.', 'danger')
            form_data = request.form.to_dict()
            form_data['custom_option_groups'] = json.loads(custom_options_json_str) if custom_options_json_str else {}
            return render_template('admin_edit_artwork.html', artwork=artwork, categories=categories, form_data=form_data)

        # --- Handle new image uploads ---
        new_image_paths = []
        if 'new_images' in request.files:
            for file in request.files.getlist('new_images'):
                if file and allowed_file(file.filename):
                    upload_result = cloudinary.uploader.upload(file)
                    new_image_paths.append(upload_result['secure_url'])
                elif file.filename != '':
                    flash(f'Invalid file type: {file.filename}', 'danger')
                    form_data = request.form.to_dict()
                    form_data['custom_option_groups'] = json.loads(custom_options_json_str) if custom_options_json_str else {}
                    return render_template('admin_edit_artwork.html', artwork=artwork, categories=categories, form_data=form_data)

        artwork.set_images_list(images_to_keep + new_image_paths)

        # --- Update custom options ---
        if custom_options_json_str:
            try:
                json.loads(custom_options_json_str)
                artwork.custom_options = custom_options_json_str
            except json.JSONDecodeError:
                flash('Invalid JSON format for custom options.', 'danger')
                form_data = request.form.to_dict()
                form_data['custom_option_groups'] = json.loads(custom_options_json_str) if custom_options_json_str else {}
                return render_template('admin_edit_artwork.html', artwork=artwork, categories=categories, form_data=form_data)
        else:
            artwork.custom_options = None

        # --- Log stock change if applicable ---
        if new_stock != old_stock:
            change_qty = new_stock - old_stock
            stock_log = StockLog(
                artwork_id=artwork.id,
                change_quantity=change_qty,
                current_stock=new_stock,
                change_type='ADD' if change_qty > 0 else 'ADJUST',
                user_id=current_user.id,
                remarks=f"Manual stock update by admin from {old_stock} → {new_stock}"
            )
            db.session.add(stock_log)

        # --- Commit everything ---
        db.session.commit()
        flash('Artwork updated successfully!', 'success')
        return redirect(url_for('admin_artworks'))

    # --- GET request: populate form_data ---
    form_data = {
        'name': artwork.name,
        'category_id': artwork.category_id,
        'is_featured': artwork.is_featured,
        'original_price': artwork.original_price,
        'cgst_percentage': artwork.cgst_percentage,
        'sgst_percentage': artwork.sgst_percentage,
        'igst_percentage': artwork.igst_percentage,
        'ugst_percentage': artwork.ugst_percentage,
        'cess_percentage': artwork.cess_percentage,
        'gst_type': artwork.gst_type,
        'stock': artwork.stock,
        'description': artwork.description,
        'shipping_charge': artwork.shipping_charge,
        'custom_option_groups': artwork.get_custom_options_dict()
    }

    return render_template('admin_edit_artwork.html', artwork=artwork, categories=categories, form_data=form_data)


@app.route('/admin/update_stock/<string:artwork_id>', methods=['POST'])
@login_required
@admin_required
def update_stock(artwork_id):
    """Admin can manually adjust artwork stock (add/reduce)."""
    artwork = db.session.get(Artwork, artwork_id)
    if not artwork:
        flash('Artwork not found.', 'danger')
        return redirect(url_for('admin_panel'))

    try:
        change_qty = int(request.form.get('change_qty', 0))
        change_type = request.form.get('change_type', 'ADJUSTMENT')
        remarks = request.form.get('remarks', '').strip()

        if change_qty == 0:
            flash('Please enter a non-zero quantity.', 'warning')
            return redirect(url_for('admin_panel'))

        # Update artwork stock
        artwork.stock += change_qty
        if artwork.stock < 0:
            artwork.stock = 0

        # Log the stock change
        stock_log = StockLog(
            artwork_id=artwork.id,
            change_quantity=change_qty,
            current_stock=artwork.stock,
            change_type=change_type.upper(),
            user_id=current_user.id,
            remarks=remarks or f"Manual {change_type.lower()} by {current_user.email}"
        )
        db.session.add(stock_log)
        db.session.commit()

        flash('Stock updated successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating stock: {e}', 'danger')

    return redirect(url_for('admin_panel'))
# In app.py, add this new route:
# In app.py, add this new route:
@app.route('/admin_panel')
@login_required
@admin_required
def admin_panel():
    """Main admin dashboard with key stats and low-stock alerts."""
    # Fetch key statistics for the dashboard tiles
    total_users = User.query.count()
    total_artworks = Artwork.query.count()
    total_orders = Order.query.count()
    pending_orders_count = Order.query.filter(Order.status.ilike('%Pending%')).count()

    # 🟡 Low-stock alert logic (threshold = 5)
    low_stock_items = Artwork.query.filter(Artwork.stock < 5).all()

    # Fetch data needed for dashboard display (e.g., recent orders, quick stock adjust)
    orders = Order.query.order_by(Order.order_date.desc()).all()
    artworks = Artwork.query.order_by(Artwork.name.asc()).all()

    return render_template(
        'admin_panel.html',
        total_users=total_users,
        total_artworks=total_artworks,
        total_orders=total_orders,
        pending_orders_count=pending_orders_count,
        orders=orders,
        artworks=artworks,
        low_stock_items=low_stock_items   # 🟢 This list powers the alert in admin_panel.html
    )


# In app.py
from flask import flash, redirect, url_for
from sqlalchemy.exc import IntegrityError  # Import IntegrityError

# NEW: Route to update artwork display order
@app.route('/admin/artwork/order/<int:artwork_id>', methods=['POST'])
@login_required
def admin_update_artwork_order(artwork_id):
    if not current_user.is_admin:
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('admin_artworks_view'))
    
    artwork = Artwork.query.get_or_404(artwork_id)
    new_order = request.form.get('display_order', type=int)

    if new_order is not None:
        artwork.display_order = new_order
        db.session.commit()
        flash('Artwork display order updated successfully!', 'success')
    
    return redirect(url_for('admin_artworks_view'))


import os
from flask import current_app
# deletes physical file also
@app.route('/admin/artwork/delete/<string:artwork_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_artwork(artwork_id):
    try:
        artwork = db.session.get(Artwork, artwork_id)
        if not artwork:
            flash('Artwork not found.', 'danger')
            return redirect(url_for('admin_artworks'))

        # Check for associated orders before attempting deletion
        has_orders = OrderItem.query.filter_by(artwork_id=artwork.id).first()
        if has_orders:
            flash("Cannot delete artwork: It is linked to existing orders. Please handle the orders first.", 'danger')
            return redirect(url_for('admin_artworks'))

        # Get the path to the uploads folder
        uploads_dir = os.path.join(current_app.root_path, 'static', 'uploads')
        
        # Split the artwork.images string into individual filenames
        image_filenames = artwork.images.split(',') if artwork.images else []

        # Delete each image file
        for filename in image_filenames:
            image_path = os.path.join(uploads_dir, filename)
            if os.path.exists(image_path):
                os.remove(image_path)
                print(f"Deleted image file: {image_path}")
        
        # Delete the artwork record from the database
        db.session.delete(artwork)
        db.session.commit()
        
        flash(f'Artwork "{artwork.name}" and its images have been deleted successfully.', 'success')
        return redirect(url_for('admin_artworks'))

    except IntegrityError as e:
        db.session.rollback()
        flash(f"Deletion failed due to a database constraint. Error: {str(e)}", 'danger')
        return redirect(url_for('admin_artworks'))
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting artwork: {e}")
        flash('An unexpected error occurred while deleting the artwork.', 'danger')
        return redirect(url_for('admin_artworks'))    


import os
from dotenv import load_dotenv
load_dotenv()


import json
import csv
import uuid
import requests
from datetime import datetime, timedelta # Ensure timedelta is imported here
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename # Added this import
from slugify import slugify

from flask import Flask, render_template, redirect, url_for, flash, request, session, jsonify, make_response, send_file
from flask import current_app
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import random
import qrcode
import io
import base64
import string 
from markupsafe import Markup

# SQLAlchemy Imports
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import Column, Integer, String, Text, Boolean, DateTime, Numeric, ForeignKey, func, or_ 
from sqlalchemy.orm import relationship
from sqlalchemy.exc import IntegrityError 

from flask_migrate import Migrate


# Email Sending
#import smtplib
#from email.mime.multipart import MIMEMultipart
#from email.mime.text import MIMEText
#from email.mime.application import MIMEApplication

# CSRF protection
from flask_wtf.csrf import CSRFProtect, generate_csrf

# PDF generation
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.lib.colors import HexColor
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    SimpleDocTemplate = SimpleDocTemplate
except ImportError:
    SimpleDocTemplate = None


























# Helper function to generate the flowables for a single invoice copy
def _get_single_invoice_flowables(order, invoice_data_safe, styles, font_name, font_name_bold):
    story_elements = []

    # Retrieve B2B fields from the safe invoice data
    is_b2b = invoice_data_safe.get('is_b2b', False)
    customer_gstin = invoice_data_safe.get('customer_gstin', '')

    # Conditionally set the invoice title
    if is_b2b:
        invoice_title = "TAX INVOICE (B2B)"
    else:
        invoice_title = "RETAIL INVOICE (B2C)"

    # Header
    story_elements.append(Paragraph(invoice_data_safe['business_name'], styles['h1']))
    story_elements.append(Paragraph(invoice_data_safe['business_address'], styles['Normal']))
    story_elements.append(Paragraph(f"GSTIN: {invoice_data_safe['gst_number']} | PAN: {invoice_data_safe['pan_number']}", styles['Normal']))
    story_elements.append(Spacer(1, 0.05 * inch))

    # Invoice Title and Details
    story_elements.append(Paragraph(invoice_title, styles['h2'])) # Use the new invoice_title variable
    
    # Use a two-column table for Invoice No and Date for better alignment
    invoice_header_data = [
        [Paragraph(f"<b>Invoice No:</b> {invoice_data_safe['invoice_number']}", styles['BoldBodyText']),
         Paragraph(f"<b>Invoice Date:</b> {invoice_data_safe['invoice_date_dt'].strftime('%d-%m-%Y')}", styles['RightAlign'])]
    ]
    invoice_header_table = Table(invoice_header_data, colWidths=[3.25*inch, 3.25*inch])
    invoice_header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    story_elements.append(invoice_header_table)
    story_elements.append(Spacer(1, 0.05 * inch))

    # --- NEW ADDRESS BLOCK ---
    # Customer Name and Phone Number
    story_elements.append(Paragraph(f"<b>Bill To:</b> {order.customer_name} ({order.customer_phone if order.customer_phone else 'N/A'})", styles['BoldBodyText']))
    story_elements.append(Spacer(1, 0.05 * inch))

    # Address Table (Shipping and Billing)
    shipping_address = order.get_shipping_address()
    shipping_address_text = (
        f"{shipping_address.get('address_line1', '')}, {shipping_address.get('address_line2', '')}, "
        f"{shipping_address.get('city', '')}, {shipping_address.get('state', '')} "
        f"<b>{shipping_address.get('pincode', '')}</b>"
    )

    # Use a two-column table for shipping and billing addresses
    address_data = [
        [
            Paragraph("<b>Shipping Address:</b>", styles['BoldBodyText']),
            Paragraph("<b>Billing Address:</b>", styles['BoldBodyText'])
        ],
        [
            Paragraph(shipping_address_text, styles['Normal']),
            Paragraph(invoice_data_safe.get('billing_address', 'N/A'), styles['Normal'])
        ]
    ]

    # Conditionally add the customer GSTIN if it's a B2B invoice
    if is_b2b and customer_gstin:
        address_data.append([Paragraph(f"Customer GSTIN:", styles['BoldBodyText']), Paragraph(customer_gstin, styles['Normal'])])
    
    address_table = Table(address_data, colWidths=[3.75*inch, 3.75*inch])
    address_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story_elements.append(address_table)
    story_elements.append(Spacer(1, 0.05 * inch))

    # Order Items Table
    data = [
        [
            Paragraph('SKU', styles['BoldBodyText']),
            Paragraph('Artwork Name', styles['BoldBodyText']),
            Paragraph('HSN Code', styles['BoldBodyText']), # NEW: HSN Code header
            Paragraph('Options', styles['BoldBodyText']),
            Paragraph('Unit Price (Excl. GST)', styles['BoldBodyText']),
            Paragraph('Qty', styles['BoldBodyText']),
            Paragraph('Taxable Value', styles['BoldBodyText']),
            Paragraph('CGST', styles['BoldBodyText']),
            Paragraph('SGST', styles['BoldBodyText']),
            Paragraph('IGST', styles['BoldBodyText']),
            Paragraph('UGST', styles['BoldBodyText']),
            Paragraph('CESS', styles['BoldBodyText']),
            Paragraph('Total (Incl. GST)', styles['BoldBodyText'])
        ]
    ]
    
    total_taxable_value = Decimal('0.00')
    total_cgst_items = Decimal('0.00')
    total_sgst_items = Decimal('0.00')
    total_igst_items = Decimal('0.00')
    total_ugst_items = Decimal('0.00')
    total_cess_items = Decimal('0.00')
    
    for item in order.items:
        options_str = ", ".join([f"{k}: {v}" for k, v in item.get_selected_options_dict().items()])
        
        # Ensure Decimal values are converted to string for display in PDF
        unit_price_excl_gst_display = f"₹{item.unit_price_before_gst:,.2f}"
        taxable_value_display = f"₹{item.total_price_before_gst:,.2f}"
        
        # Conditional display for GST components
        cgst_display = f"₹{item.cgst_amount:,.2f} ({item.cgst_percentage_applied}%)" if item.cgst_amount > Decimal('0.00') else ""
        sgst_display = f"₹{item.sgst_amount:,.2f} ({item.sgst_percentage_applied}%)" if item.sgst_amount > Decimal('0.00') else ""
        igst_display = f"₹{item.igst_amount:,.2f} ({item.igst_percentage_applied}%)" if item.igst_amount > Decimal('0.00') else ""
        ugst_display = f"₹{item.ugst_amount:,.2f} ({item.ugst_percentage_applied}%)" if item.ugst_amount > Decimal('0.00') else ""
        cess_display = f"₹{item.cess_amount:,.2f} ({item.cess_percentage_applied}%)" if item.cess_amount > Decimal('0.00') else ""
        
        total_incl_gst_display = f"₹{item.total_price_incl_gst:,.2f}"

        data.append([
            Paragraph(item.artwork.sku, styles['TableCell']),
            Paragraph(item.artwork.name, styles['TableCellLeft']),
            Paragraph(f"{item.artwork.hsn_code or ''}<br/><font size=\"8\">{item.artwork.hsn_description or ''}</font>", styles['TableCell']),
            Paragraph(options_str, styles['TableCellLeft']),
            Paragraph(unit_price_excl_gst_display, styles['TableCellRight']),
            Paragraph(str(item.quantity), styles['TableCell']),
            Paragraph(taxable_value_display, styles['TableCellRight']),
            Paragraph(cgst_display, styles['TableCellRight']),
            Paragraph(sgst_display, styles['TableCellRight']),
            Paragraph(igst_display, styles['TableCellRight']),
            Paragraph(ugst_display, styles['TableCellRight']),
            Paragraph(cess_display, styles['TableCellRight']),
            Paragraph(total_incl_gst_display, styles['TableCellRight'])
        ])
        total_taxable_value += item.total_price_before_gst
        total_cgst_items += item.cgst_amount
        total_sgst_items += item.sgst_amount
        total_igst_items += item.igst_amount
        total_ugst_items += item.ugst_amount
        total_cess_items += item.cess_amount

    col_widths = [
        0.5*inch, # SKU
        0.8*inch, # Artwork Name
        0.7*inch, # HSN Code (NEW)
        0.7*inch, # Options
        0.6*inch, # Unit Price (Excl. GST)
        0.3*inch, # Qty
        0.8*inch, # Taxable Value
        0.6*inch, # CGST
        0.6*inch, # SGST
        0.6*inch, # IGST
        0.6*inch, # UGST
        0.5*inch, # CESS
        0.8*inch  # Total (Incl. GST)
    ]
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#FFBF00')), # Golden Saffron header
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'), # Align artwork name left
        ('ALIGN', (2, 0), (2, -1), 'CENTER'), # NEW: Align HSN Code center
        ('ALIGN', (3, 0), (3, -1), 'LEFT'), # Align options left
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'), # Align unit price right
        ('ALIGN', (5, 0), (5, -1), 'CENTER'), # Align quantity center
        ('ALIGN', (6, 0), (-1, -1), 'RIGHT'), # Align Taxable Value, GST and total amounts right
        ('FONTNAME', (0, 0), (-1, 0), font_name_bold), # Use registered bold font
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4), # Reduced padding
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#FAF9F6')), # Light background for rows
        ('GRID', (0, 0), (-1, -1), 0.25, HexColor('#E5E7EB')), # Lighter grid lines
        ('BOX', (0, 0), (-1, -1), 0.25, HexColor('#E5E7EB')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 2), # Further reduced padding
        ('RIGHTPADDING', (0,0), (-1,-1), 2), # Further reduced padding
        ('TOPPADDING', (0,0), (-1,-1), 2), # Further reduced padding
        ('BOTTOMPADDING', (0,0), (-1,-1), 2), # Further reduced padding
    ]))
    story_elements.append(table)
    story_elements.append(Spacer(1, 0.05 * inch)) # Reduced space
    # Summary Table
    summary_data = []
    # REFINED: Use the pre-calculated total_taxable_value for clarity
    summary_data.append([Paragraph('Subtotal (Excl. GST):', styles['RightAlign']), Paragraph(f"₹{total_taxable_value:,.2f}", styles['RightAlign'])])
    if total_cgst_items > Decimal('0.00'):
        summary_data.append([Paragraph('CGST:', styles['RightAlign']), Paragraph(f"₹{total_cgst_items:,.2f}", styles['RightAlign'])])
    if total_sgst_items > Decimal('0.00'):
        summary_data.append([Paragraph('SGST:', styles['RightAlign']), Paragraph(f"₹{total_sgst_items:,.2f}", styles['RightAlign'])])
    if total_igst_items > Decimal('0.00'):
        summary_data.append([Paragraph('IGST:', styles['RightAlign']), Paragraph(f"₹{total_igst_items:,.2f}", styles['RightAlign'])])
    if total_ugst_items > Decimal('0.00'):
        summary_data.append([Paragraph('UGST:', styles['RightAlign']), Paragraph(f"₹{total_ugst_items:,.2f}", styles['RightAlign'])])
    if total_cess_items > Decimal('0.00'):
        summary_data.append([Paragraph('CESS:', styles['RightAlign']), Paragraph(f"₹{total_cess_items:,.2f}", styles['RightAlign'])])
    
    summary_data.append([Paragraph('Shipping Charge:', styles['RightAlign']), Paragraph(f"₹{order.shipping_charge:,.2f}", styles['RightAlign'])])
    summary_data.append([Paragraph('<b>Grand Total (Incl. GST):</b>', styles['BoldBodyText']), Paragraph(f"<b>₹{order.total_amount:,.2f}</b>", styles['BoldBodyText'])])
    
    summary_table = Table(summary_data, colWidths=[5*inch, 2.5*inch]) # Adjusted total width to 7.5 inches
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), font_name_bold), # Bold for Grand Total
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2), # Reduced padding
        ('GRID', (0, 0), (-1, -1), 0.25, HexColor('#E5E7EB')),
        ('BOX', (0, 0), (-1, -1), 0.25, HexColor('#E5E7EB')),
    ]))
    story_elements.append(summary_table)
    story_elements.append(Spacer(1, 0.1 * inch)) # Reduced space

    # WhatsApp Message
    story_elements.append(Paragraph("Please WhatsApp to +919123700057 for easy returns and refund.", styles['Normal']))
    story_elements.append(Spacer(1, 0.1 * inch)) # Reduced space

    # Authorized Signatory and Office Seal Block
    story_elements.append(Spacer(1, 0.2 * inch)) # Add some space before the signature block

    signature_data = [
        [Paragraph("<b>For Karthika Futures</b>", styles['RightAlign'])],
        [Spacer(1, 0.5 * inch)], # Space for signature
        [Paragraph("_________________________", styles['RightAlign'])], # Corrected line
        [Paragraph("<b>Authorized Signatory</b>", styles['RightAlign'])],
        [Spacer(1, 0.1 * inch)], # Space between signatory and seal area
        [Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['RightAlign'])], # Date for ink seal
        [Paragraph("(Office Seal)", styles['RightAlign'])] # Placeholder for office seal
    ]

    signature_table = Table(signature_data, colWidths=[7.5*inch]) # Table spans full content width
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('ALIGN', (0,2), (0,2), 'RIGHT'),
        ('ALIGN', (0,3), (0,3), 'RIGHT'),
        ('ALIGN', (0,5), (0,5), 'RIGHT'),
        ('ALIGN', (0,6), (0,6), 'RIGHT'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    story_elements.append(signature_table)

    story_elements.append(Spacer(1, 0.1 * inch)) # Space before footer

    # Footer
    story_elements.append(Paragraph("Thank you for your purchase and devotion!", styles['Normal']))
    story_elements.append(Paragraph(f"Invoice Status: <b>{invoice_data_safe['invoice_status']}</b>", styles['BoldBodyText']))
    story_elements.append(Paragraph("This is a computer-generated invoice and does not require a signature.", styles['Normal']))
    story_elements.append(Spacer(1, 0.05 * inch)) # Reduced space
    story_elements.append(Paragraph(f"Contact: {order.customer_phone if order.customer_phone else 'N/A'} | {order.customer_email if order.customer_email else 'N/A'}", styles['Footer']))

    return story_elements

@app.route('/admin/order/<order_id>/generate_and_upload_invoice', methods=['POST'])
@admin_required
def admin_generate_and_upload_invoice_pdf(order_id):
    """
    Admin route: Generates and uploads the invoice PDF to Cloudinary (public access).
    """
    order = db.session.get(Order, order_id)
    if not order:
        flash('Order not found.', 'danger')
        return redirect(url_for('admin_dashboard'))

    pdf_buffer = generate_invoice_pdf_buffer(order)
    if pdf_buffer is None:
        flash('PDF generation failed.', 'danger')
        return redirect(url_for('admin_edit_invoice', order_id=order_id))

    try:
        pdf_buffer.seek(0)
        folder_name = current_app.config.get("CLOUDINARY_INVOICE_FOLDER", "invoices")
        public_id = f"invoice_{order.id}"

        # ✅ Upload with public access
        upload_result = cloudinary.uploader.upload(
            pdf_buffer,
            resource_type="raw",
            type="upload",
            folder=folder_name,
            public_id=public_id,
            overwrite=True,
            format="pdf",
            access_mode="public"   # ✅ makes PDF link public
        )

        # ✅ Save the secure URL directly (it will be versioned but fully accessible)
        file_url = upload_result.get("secure_url")
        order.invoice_file_url = file_url
        db.session.commit()

        flash(f"Invoice uploaded successfully. Shareable URL: {order.invoice_file_url}", "success")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Cloudinary upload error for Order {order_id}: {e}", exc_info=True)
        flash(f"Error uploading invoice: {e}", "danger")

    return redirect(url_for('admin_edit_invoice', order_id=order_id))


@app.route('/admin/edit-invoice/<order_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_invoice(order_id):
    order = Order.query.get_or_404(order_id)
    
    if request.method == 'POST':
        # Collect all invoice details from the form
        invoice_details = {
            'business_name': request.form.get('business_name'),
            'gst_number': request.form.get('gst_number'),
            'pan_number': request.form.get('pan_number'),
            'business_address': request.form.get('business_address'),
            'invoice_number': request.form.get('invoice_number'),
            'invoice_date': request.form.get('invoice_date'),
            'billing_address': request.form.get('billing_address'), # This field is not in the HTML, but kept for completeness if you add it
            # Convert Decimal types to string before storing in JSON
            'gst_rate_applied': str(Decimal(request.form.get('gst_rate', '0.00'))), 
            'shipping_charge': str(Decimal(request.form.get('shipping_charge', '0.00'))), 
            'final_invoice_amount': str(Decimal(request.form.get('total_amount', '0.00'))), 
            'invoice_status': request.form.get('invoice_status', 'Generated'), # Default if not explicitly set
            'is_held_by_admin': 'is_held_by_admin' in request.form, # Checkbox
            'cgst_amount': str(Decimal(request.form.get('cgst_amount', '0.00'))), 
            'sgst_amount': str(Decimal(request.form.get('sgst_amount', '0.00'))),
            'igst_amount': str(Decimal(request.form.get('igst_amount', '0.00'))),
            'ugst_amount': str(Decimal(request.form.get('ugst_amount', '0.00'))),
            'cess_amount': str(Decimal(request.form.get('cess_amount', '0.00'))), # NEW
            # NEW: Add B2B fields to the dictionary
            'is_b2b': 'is_b2b' in request.form,
            'customer_gstin': request.form.get('customer_gstin', '').upper()
        }
        
        order.set_invoice_details(invoice_details)
        order.shipping_charge = Decimal(invoice_details['shipping_charge']) 
        order.total_amount = Decimal(invoice_details['final_invoice_amount']) 
        db.session.commit()
        flash('Invoice details updated successfully!', 'success')
        return redirect(url_for('admin_dashboard')) # Or admin_orders_view if you create one
    
    # For GET request, parse invoice_details and convert date strings back to datetime objects
    invoice_data = order.get_invoice_details()
    invoice_data['customer_gstin'] = invoice_data.get('customer_gstin') or order.customer_gstin or ''
    if 'invoice_date' in invoice_data and invoice_data['invoice_date']:
        try:
            invoice_data['invoice_date'] = datetime.strptime(invoice_data['invoice_date'], '%Y-%m-%d')
        except ValueError:
            invoice_data['invoice_date'] = None # Handle potential format issues
    if 'due_date' in invoice_data and invoice_data['due_date']:
        try:
            invoice_data['due_date'] = datetime.strptime(invoice_data['due_date'], '%Y-%m-%d')
        except ValueError:
            invoice_data['due_date'] = None # Handle potential format issues

    return render_template('admin_edit_invoice.html', order=order, invoice_data=invoice_data)

@app.route("/generate_invoice_pdf/<order_id>", methods=["GET", "POST"])
@login_required
@admin_required
def generate_invoice_pdf(order_id):
    """
    Admin route to generate an invoice PDF, preview (GET), or upload to Cloudinary (POST).
    """
    order = db.session.get(Order, order_id)
    if not order:
        flash("Order not found.", "danger")
        return redirect(url_for("admin_dashboard"))

    # CASE 1: GET → return PDF in browser
    if request.method == "GET":
        try:
            pdf_buffer = generate_invoice_pdf_buffer(order)
            if not pdf_buffer:
                flash("Failed to generate invoice PDF.", "danger")
                return redirect(url_for("admin_edit_invoice", order_id=order.id))

            return send_file(
                pdf_buffer,
                as_attachment=False,
                download_name=f"invoice_{order.id}.pdf",
                mimetype="application/pdf",
            )
        except Exception as e:
            app.logger.error(f"Error generating invoice PDF for {order_id}: {e}")
            flash("Error generating PDF.", "danger")
            return redirect(url_for("admin_edit_invoice", order_id=order.id))

    # CASE 2: POST → upload to Cloudinary
    if request.method == "POST":
        try:
            pdf_buffer = generate_invoice_pdf_buffer(order)
            if not pdf_buffer:
                return jsonify(success=False, message="Failed to generate invoice."), 500

            pdf_buffer.seek(0)
            folder_name = current_app.config.get("CLOUDINARY_INVOICE_FOLDER", "invoices")
            public_id = f"invoice_{order.id}"

            # ✅ Upload with public access mode
            upload_result = cloudinary.uploader.upload(
                pdf_buffer,
                resource_type="raw",
                folder=folder_name,
                public_id=public_id,
                overwrite=True,
                format="pdf",
                access_mode="public"   # ✅ public link fix
            )

            file_url = upload_result.get("secure_url")
            order.invoice_file_url = file_url
            db.session.commit()

            return jsonify(success=True, message="Invoice generated & uploaded.", file_url=file_url)

        except Exception as e:
            app.logger.error(f"Error in POST /generate_invoice_pdf/{order_id}: {e}")
            return jsonify(success=False, message="Server error while generating invoice."), 500

def generate_invoice_pdf_buffer(order):
    if SimpleDocTemplate is None:
        return None

    # Get stored invoice data
    invoice_data = order.get_invoice_details()

    # ✅ Determine B2B or B2C automatically
    is_b2b = bool(order.customer_gstin and order.customer_gstin.strip())
    invoice_type = "B2B" if is_b2b else "B2C"

    # ✅ Prepare safe data
    invoice_data_safe = {
        'business_name': invoice_data.get('business_name') or app.config['OUR_BUSINESS_NAME'],
        'gst_number': invoice_data.get('gst_number') or app.config['OUR_GSTIN'],
        'pan_number': invoice_data.get('pan_number') or app.config['OUR_PAN'],
        'business_address': invoice_data.get('business_address') or app.config['OUR_BUSINESS_ADDRESS'],
        'invoice_number': invoice_data.get('invoice_number') or order.id,
        'invoice_date': invoice_data.get('invoice_date') or datetime.utcnow().strftime('%Y-%m-%d'),
        'billing_address': invoice_data.get('billing_address') or ', '.join(filter(None, [
            order.customer_name,
            order.customer_phone,
            order.get_shipping_address().get('address_line1'),
            order.get_shipping_address().get('address_line2'),
            order.get_shipping_address().get('city'),
            order.get_shipping_address().get('state'),
            order.get_shipping_address().get('pincode')
        ])),
        'gst_rate_applied': Decimal(invoice_data.get('gst_rate_applied', '0.00')), 
        'shipping_charge': Decimal(invoice_data.get('shipping_charge', '0.00')), 
        'final_invoice_amount': Decimal(invoice_data.get('final_invoice_amount', '0.00')), 
        'invoice_status': invoice_data.get('invoice_status', 'Generated'),
        'cgst_amount': Decimal(invoice_data.get('cgst_amount', '0.00')), 
        'sgst_amount': Decimal(invoice_data.get('sgst_amount', '0.00')),
        'igst_amount': Decimal(invoice_data.get('igst_amount', '0.00')),
        'ugst_amount': Decimal(invoice_data.get('ugst_amount', '0.00')),
        'cess_amount': Decimal(invoice_data.get('cess_amount', '0.00')),
        'is_b2b': is_b2b,
        'customer_gstin': order.customer_gstin or '',
        'invoice_type': invoice_type
    }

    # ✅ Convert invoice_date safely
    if 'invoice_date' in invoice_data_safe and invoice_data_safe['invoice_date']:
        if isinstance(invoice_data_safe['invoice_date'], str):
            try:
                invoice_data_safe['invoice_date_dt'] = datetime.strptime(invoice_data_safe['invoice_date'], '%Y-%m-%d')
            except ValueError:
                invoice_data_safe['invoice_date_dt'] = datetime.utcnow()
        else:
            invoice_data_safe['invoice_date_dt'] = invoice_data_safe['invoice_date']
    else:
        invoice_data_safe['invoice_date_dt'] = datetime.utcnow()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=0.3*inch, rightMargin=0.3*inch,
                            topMargin=0.3*inch, bottomMargin=0.3*inch)

    # ✅ Fonts setup
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf'))
        font_name = 'DejaVuSans'
        font_name_bold = 'DejaVuSans-Bold'
    except Exception:
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'

    # ✅ Styles setup
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = font_name
    styles['Normal'].fontSize = 8
    styles['Normal'].leading = 9
    styles['Normal'].alignment = TA_LEFT
    styles['Normal'].spaceAfter = 1

    styles['h1'].fontName = font_name_bold
    styles['h1'].fontSize = 14
    styles['h1'].leading = 16
    styles['h1'].alignment = TA_CENTER
    styles['h1'].spaceAfter = 3

    styles['h2'].fontName = font_name_bold
    styles['h2'].fontSize = 12
    styles['h2'].leading = 14
    styles['h2'].alignment = TA_LEFT
    styles['h2'].spaceAfter = 4

    if 'RightAlign' not in styles:
        styles.add(ParagraphStyle(name='RightAlign', parent=styles['Normal']))
    styles['RightAlign'].alignment = TA_RIGHT

    if 'BoldBodyText' not in styles:
        styles.add(ParagraphStyle(name='BoldBodyText', parent=styles['Normal']))
    styles['BoldBodyText'].fontName = font_name_bold

    if 'Footer' not in styles:
        styles.add(ParagraphStyle(name='Footer', parent=styles['Normal']))
    styles['Footer'].fontSize = 6
    styles['Footer'].alignment = TA_CENTER

    if 'TableCell' not in styles:
        styles.add(ParagraphStyle(name='TableCell', parent=styles['Normal']))
    styles['TableCell'].alignment = TA_CENTER
    styles['TableCell'].fontSize = 7

    if 'TableCellLeft' not in styles:
        styles.add(ParagraphStyle(name='TableCellLeft', parent=styles['Normal']))
    styles['TableCellLeft'].alignment = TA_LEFT

    if 'TableCellRight' not in styles:
        styles.add(ParagraphStyle(name='TableCellRight', parent=styles['Normal']))
    styles['TableCellRight'].alignment = TA_RIGHT

    # ✅ Build PDF
    story = []

    # Add invoice type visibly
    story.append(Paragraph(f"<b>Invoice Type:</b> {invoice_type}", styles["BodyText"]))
    story.append(Spacer(1, 4))

    story += _get_single_invoice_flowables(order, invoice_data_safe, styles, font_name, font_name_bold)
    story.append(Spacer(1, 0.5 * inch))
    story += _get_single_invoice_flowables(order, invoice_data_safe, styles, font_name, font_name_bold)

    doc.build(story)
    buffer.seek(0)
    return buffer




@app.route('/admin/order/<order_id>/remove_invoice_pdf', methods=['POST'])
@admin_required
def admin_remove_invoice_pdf(order_id):
    order = db.session.get(Order, order_id)
    if not order:
        flash('Order not found.', 'danger')
        return redirect(url_for('admin_dashboard'))

    # Clear the URL in the database
    order.invoice_file_url = None
    db.session.commit()
    
    flash(f'Invoice link removed for Order ID {order_id}. The customer will no longer see the download option.', 'warning')
    return redirect(url_for('admin_edit_invoice', order_id=order_id))


from flask import redirect, flash, url_for
from flask_login import login_required, current_user

import os
import re # <-- Ensure this is at the top of your app.py along with other imports

# Assuming INVOICE_RESOURCE_TYPE = "raw" is defined globally

# REMOVED: The problematic generate_signed_invoice_url function is no longer needed.
# We consolidate the logic into the route below for robustness.

# --- FIX: The main download route now handles URL signing directly ---
@app.route('/order/download/<order_id>', methods=['GET'])
@login_required
def order_download_invoice(order_id):
    import requests
    from flask import make_response
    
    order = db.session.get(Order, order_id)
    if not order or not order.invoice_file_url:
        flash("Invoice not found or not yet generated.", 'danger')
        return redirect(url_for('user_orders'))

    try:
        # We need the full versioned path for the signed URL
        # We must redefine the required global constant here to ensure it's available
        INVOICE_RESOURCE_TYPE = "raw" 
        
        # Extract the full path including the version number (vXXXX/folder/filename)
        # This is the correct regex needed to resolve the 404 error
        match = re.search(r'/(?:raw|image|video)/upload/(.+?)(?:\.\w+)?$', order.invoice_file_url)
        
        if not match:
             raise ValueError(f"Could not extract versioned Public ID from URL: {order.invoice_file_url}")
        
        public_id_for_signing = match.group(1)
        
        # 1. Generate the secure, signed URL directly
        # We use type="upload" as it is the standard and most robust configuration for signed URLs
        final_download_url, options = cloudinary.utils.cloudinary_url(
            public_id_for_signing, 
            resource_type=INVOICE_RESOURCE_TYPE,
            type="upload",  # <-- Using 'upload' is the most common correct configuration for signing public files
            format="pdf",
            sign_url=True,
            attachment=f"Invoice_{order.id}.pdf",
            expires_at=int((datetime.now() + timedelta(minutes=10)).timestamp()) 
        )
        
        # 2. Server-side fetch (using requests)
        response = requests.get(final_download_url, timeout=60)
        
        # Check for 404/401/403 errors
        if response.status_code != 200:
             # Include the actual Cloudinary error message in the flash message for debugging
            error_details = response.text if response.text else response.reason
            raise requests.exceptions.HTTPError(
                f"{response.status_code} Client Error: {error_details} for url: {final_download_url}"
            )
        
        # Create a Flask response to stream the file to the user
        response = make_response(response.content)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=Invoice_{order.id}.pdf'
        return response

    except requests.exceptions.RequestException as e:
        current_app.logger.error(f"Error fetching invoice PDF for {order_id}: {e}")
        flash(f"Failed to download the invoice. A server error occurred: {e}", 'danger')
    except Exception as e:
        current_app.logger.error(f"General error in invoice download for {order_id}: {e}")
        flash(f"An unexpected error occurred during download. {e}", 'danger')
        
    return redirect(url_for('user_orders'))




@app.route('/download-invoice/<order_id>')
@login_required 
def download_invoice(order_id):
    # This route is simplified to redirect to the client-side mechanism
    return redirect(url_for('order_download_invoice', order_id=order_id))

@app.route('/order/<string:order_id>/generate-invoice')
@login_required
def user_generate_invoice(order_id):
    """
    Allows a logged-in user to regenerate and download their invoice PDF
    only after the order is shipped or delivered.
    """
    order = db.session.get(Order, order_id)
    if not order:
        flash("Order not found.", "danger")
        return redirect(url_for("user_orders"))

    # ✅ Only allow after shipment
    if str(order.user_id) != str(current_user.id):
        flash("Unauthorized access.", "danger")
        return redirect(url_for("user_orders"))

    if order.status not in ["Shipped", "Delivered"]:
        flash("Invoice will be available once your order is shipped.", "warning")
        return redirect(url_for("user_orders"))

    # ✅ Generate PDF now
    pdf_buffer = generate_invoice_pdf_buffer(order)
    if not pdf_buffer:
        flash("Could not generate invoice.", "danger")
        return redirect(url_for("user_orders"))

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"invoice_{order.id}.pdf",
        mimetype="application/pdf"
    )





















@app.route('/about')
def about():
    return render_template('about.html')



# --- User Profile & Orders ---
@app.route('/user-profile')
@login_required
def user_profile():
    user_addresses = current_user.addresses
    return render_template('user_profile.html', user_addresses=user_addresses)

@app.route('/add-address-form', methods=['GET'])
@login_required
def add_address_form():
    return render_template('add_address_form.html')


@app.route('/add-address', methods=['POST'])
@login_required
def add_address():
    full_name = request.form.get('full_name')
    phone = request.form.get('phone')
    address_line1 = request.form.get('address_line1')
    address_line2 = request.form.get('address_line2')
    city = request.form.get('city')
    state = request.form.get('state')
    pincode = request.form.get('pincode')
    is_default = 'is_default' in request.form

    if not all([full_name, phone, address_line1, city, state, pincode]):
        flash('Please fill in all required address fields.', 'danger')
        return redirect(url_for('purchase_form')) # Corrected redirect for validation failure

    if is_default:
        # Unset previous default address
        for addr in current_user.addresses:
            if addr.is_default:
                addr.is_default = False

    new_address = Address(
        user_id=current_user.id,
        full_name=full_name,
        phone=phone,
        address_line1=address_line1,
        address_line2=address_line2,
        city=city,
        state=state,
        pincode=pincode,
        is_default=is_default
    )
    if new_address.is_default:
        # Minimal change: unset any other default addresses for this user
        Address.query.filter_by(user_id=current_user.id, is_default=True).update({"is_default": False})
    db.session.add(new_address)
    db.session.commit()
    flash('Address added successfully!', 'success')
    return redirect(url_for('purchase_form')) # Corrected redirect for success


@app.route('/my-addresses')
@login_required
def my_addresses():
    addresses = Address.query.filter_by(user_id=current_user.id).order_by(Address.is_default.desc(), Address.id.asc()).all()
    return render_template('my_addresses.html', addresses=addresses)


@app.route('/admin/screenshots')
@login_required
@admin_required
def view_uploaded_screenshots():
    orders = Order.query.filter(Order.payment_screenshot.isnot(None)).order_by(Order.order_date.desc()).all()
    missing_files = []

    for order in orders:
        path = os.path.join('static', order.payment_screenshot)
        if not os.path.exists(path):
            missing_files.append(order.order_id)

    return render_template('admin_screenshots.html', orders=orders, missing_files=missing_files)

@app.route('/edit-address/<address_id>', methods=['GET', 'POST'])
@login_required
def edit_address(address_id):
    # Fetch address belonging to current user only
    address = Address.query.filter_by(id=address_id, user_id=current_user.id).first_or_404()

    if request.method == 'POST':
        address.label = request.form.get('label', '')
        address.full_name = request.form.get('full_name', '')
        address.phone = request.form.get('phone', '')
        address.address_line1 = request.form.get('address_line1', '')
        address.address_line2 = request.form.get('address_line2', '')
        address.pincode = request.form.get('pincode', '')
        address.city = request.form.get('city', '')
        address.state = request.form.get('state', '')
        address.is_default = 'is_default' in request.form

        # If user set this as default, unset all others
        if address.is_default:
            other_addresses = Address.query.filter(Address.user_id == current_user.id, Address.id != address.id)
            for addr in other_addresses:
                addr.is_default = False

        db.session.commit()
        flash("Address updated successfully!", "success")
        return redirect(url_for('my_addresses'))  # Or wherever you show the list of addresses

    return render_template("edit_address.html", address=address)



@app.route('/delete-address/<address_id>', methods=['GET', 'POST'])
@login_required
def delete_address(address_id):
    if request.method == 'GET':
        flash('Invalid request method for deleting address.', 'warning')
        return redirect(url_for('user_profile'))

    address = db.session.get(Address, address_id)
    if not address:
        flash('Address not found.', 'danger')
        return redirect(url_for('user_profile'))

    if address.user_id != current_user.id:
        flash('You are not authorized to delete this address.', 'danger')
        return redirect(url_for('user_profile'))

    if address.is_default and len(current_user.addresses) > 1:
        flash('Cannot delete default address if other addresses exist. Please set another address as default first.', 'danger')
        return redirect(url_for('user_profile'))

    if db.session.query(Order).filter_by(shipping_address_id=address_id).count() > 0:
        flash('Cannot delete address: it is linked to existing orders.', 'danger')
        return redirect(url_for('user_profile'))

    db.session.delete(address)
    db.session.commit()
    flash('Address deleted successfully!', 'success')
    return redirect(url_for('user_profile'))


@app.route('/user-orders')
@login_required
def user_orders():
    if current_user.is_admin():
        orders = Order.query.order_by(Order.order_date.desc()).all()
    else:
        orders = Order.query.filter_by(user_id=current_user.id).order_by(Order.order_date.desc()).all()
    return render_template('user_orders.html', orders=orders)


import os
from werkzeug.utils import secure_filename
from flask import current_app

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/user/order/<order_id>/edit_screenshot', methods=['POST'])
@login_required
@csrf.exempt  # If you're using AJAX or ensure CSRF token included
def edit_payment_screenshot(order_id):
    # Allow admin to edit any order, user can edit only their own
    if current_user.is_admin():
        order = Order.query.filter_by(id=order_id).first()
    else:
        order = Order.query.filter_by(id=order_id, user_id=current_user.id).first()

    if not order:
        print(f"DEBUG: Order NOT found for ID={order_id}.")
        flash('Order not found or you are not authorized to view it.', 'danger')
        return redirect(url_for('user_orders'))

    if 'screenshot' not in request.files:
        flash("No screenshot uploaded.", "warning")
        return redirect(url_for('user_orders'))

    file = request.files['screenshot']
    if file and allowed_file(file.filename):
        filename = secure_filename(f"{order.id}_screenshot.{file.filename.rsplit('.', 1)[1].lower()}")
        filepath = os.path.join('static', 'payment_screenshots', filename)

        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        # Delete old screenshot if exists
        if order.payment_screenshot:
            try:
                old_screenshot_path = os.path.join('static', order.payment_screenshot)
                if os.path.exists(old_screenshot_path):
                    os.remove(old_screenshot_path)
            except Exception as e:
                print(f"Warning: Could not delete old screenshot: {e}")

        file.save(filepath)
        order.payment_screenshot = f'payment_screenshots/{filename}'
        db.session.commit()

        flash("Screenshot uploaded successfully.", "success")
    else:
        flash("Choose an image first & then Upload Screenshot. Only PNG, JPG, JPEG allowed.", "danger")

    return redirect(url_for('user_orders'))


from flask import jsonify, request

@app.route('/delete-order/<order_id>', methods=['POST', 'DELETE'])
@login_required
def delete_order(order_id):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    order = db.session.get(Order, order_id)
    if not order:
        if is_ajax:
            return jsonify(success=False, message='Order not found.'), 404
        flash('Order not found.', 'danger')
        return redirect(url_for('user_orders'))

    if not current_user.is_admin() and order.status not in ['Pending Payment', 'Payment Failed', 'Cancelled by User', 'Cancelled by Admin']:
        if is_ajax:
            return jsonify(success=False, message='You are not allowed to delete this order.'), 403
        flash('This order cannot be deleted.', 'danger')
        return redirect(url_for('user_orders'))

    try:
        for item in order.items:
            artwork = db.session.get(Artwork, item.artwork_id)
            if artwork:
                artwork.stock += item.quantity

        db.session.delete(order)
        db.session.commit()
        if is_ajax:
            return jsonify(success=True, message='Order deleted successfully.')
        flash('Order deleted successfully.', 'success')

    except Exception as e:
        db.session.rollback()
        if is_ajax:
            return jsonify(success=False, message=f'Error deleting order: {e}'), 500
        flash(f'Error deleting order: {e}', 'danger')

    return redirect(url_for('admin_orders_view' if current_user.is_admin() else 'user_orders'))


# --- Initialize DB and Migrate Data on First Run ---
with app.app_context():
    db.create_all()

    # Create default admin user if not exists
    default_admin_email = os.environ.get('DEFAULT_ADMIN_EMAIL', 'admin@karthikafutures.com')
    default_admin_password = os.environ.get('DEFAULT_ADMIN_PASSWORD', 'adminpass') # Change this in production!

    admin_user_exists = User.query.filter_by(email=default_admin_email).first()
    if not admin_user_exists:
        initial_admin = User(
            email=default_admin_email,
            full_name="Karthika Futures Admin",
            role='admin',
            email_verified=True # Admin email is pre-verified
        )
        initial_admin.set_password(default_admin_password)
        db.session.add(initial_admin)
        db.session.commit()

        # Create a default address for the admin user
        initial_admin_address = Address(
            user_id=initial_admin.id,
            label="Admin Office",
            full_name="Karthika Futures Admin",
            phone="9999999999",
            address_line1="Admin Office, 123 Main St",
            address_line2="",
            city="Spiritual City",
            state="Karnataka",
            pincode="560001",
            is_default=True
        )
        db.session.add(initial_admin_address)
        db.session.commit()
        print(f"Default admin '{default_admin_email}' created.")
        print(f"\n--- IMPORTANT: DEFAULT ADMIN CREATED ---")
        print(f"Email: {default_admin_email}")
        print(f"Password: {default_admin_password}")
        print(f"Login at: http://127.0.0.1:5000/admin-login")
        print(f"----------------------------------------\n")
    else:
        # Check if existing user with default admin email is actually an admin
        if not admin_user_exists.is_admin(): # Changed to call is_admin()
            print(f"\n--- WARNING: Account '{default_admin_email}' exists but is not admin. ---")
            print(f"Please manually change role to 'admin' for user {default_admin_email} in the database if needed.")
            print(f"------------------------------------------------------------------\n")
        else:
            print("Admin user already exists. Skipping default admin creation.")

    # NEW: Initialize order_id_sequence if it doesn't exist
    order_sequence_exists = db.session.get(SequenceCounter, 'order_id_sequence')
    if not order_sequence_exists:
        initial_sequence = SequenceCounter(id='order_id_sequence', current_value=15416760)
        db.session.add(initial_sequence)
        db.session.commit()
        print("Initialized 'order_id_sequence' to 15416760.")
    
    # NEW: Check for CESS column in Artwork and OrderItem models
    # This is a simple check. For production, use Flask-Migrate.
    from sqlalchemy import inspect
    inspector = inspect(db.engine)
    artwork_columns = [col['name'] for col in inspector.get_columns('artwork')]
    order_item_columns = [col['name'] for col in inspector.get_columns('order_item')]

    if 'cess_percentage' not in artwork_columns or \
       'cess_percentage_applied' not in order_item_columns or \
       'cess_amount' not in order_item_columns:
        # Removed the print statement for this warning
        pass # Keep this for schema check, but no print output

# Ecommerce update

#@app.route("/check-db")
#def check_db():
    #uri = app.config.get("SQLALCHEMY_DATABASE_URI", "Not Set")
    #return f"Database URI in use: {uri}"

@app.route("/version")
def version():
    return "Karthika Futures | Build: 2025-07-28 10:45 AM"

@app.route('/admin/console', methods=['GET'])
@login_required
@admin_required
def admin_console():
    import traceback
    from datetime import datetime, timedelta
    from decimal import Decimal

    try:
        # --- Base query: newest first ---
        query = Order.query.order_by(Order.order_date.desc())

        # ---- Read filters from URL ----
        business_type = request.args.get('business_type', 'all')
        status       = request.args.get('status', 'all')
        start_date   = request.args.get('start_date', '')  # 'YYYY-MM-DD'
        end_date     = request.args.get('end_date', '')    # 'YYYY-MM-DD'
        search       = request.args.get('search', '').strip()

        # ---- B2B / B2C based on GSTIN ----
        if business_type == 'b2b':
            query = query.filter(
                Order.customer_gstin.isnot(None),
                Order.customer_gstin != ''
            )
        elif business_type == 'b2c':
            query = query.filter(
                db.or_(Order.customer_gstin.is_(None), Order.customer_gstin == '')
            )

        # ---- Status filter ----
        if status and status != 'all':
            query = query.filter(Order.status == status)

        # ---- Date range filter (order_date) ----
        try:
            if start_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(Order.order_date >= start_dt)
            if end_date:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(Order.order_date < end_dt)
        except ValueError:
            flash("Invalid date format for filtering.", "warning")

        # ---- Search (Order ID / customer name / email / phone / pincode / city) ----
        if search:
            like = f"%{search}%"
            query = (query
                     .join(User)
                     .outerjoin(Address, Order.shipping_address_id == Address.id)
                     .filter(
                         db.or_(
                             Order.id.ilike(like),
                             User.full_name.ilike(like),
                             User.email.ilike(like),
                             User.phone.ilike(like),
                             Address.pincode.ilike(like),
                             Address.city.ilike(like),
                         )
                     ))

        # ---- Pagination ----
        page = request.args.get('page', 1, type=int)
        per_page = 25
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        # ---- Summary stats (all orders) ----
        total_orders = Order.query.count()
        today_orders = Order.query.filter(
            db.func.date(Order.order_date) == db.func.current_date()
        ).count()
        pending_orders = Order.query.filter_by(status='Pending Payment').count()
        total_revenue = (db.session.query(db.func.sum(Order.total_amount))
                         .scalar() or Decimal('0.00'))

        # Status distribution for small chart
        status_rows = (db.session.query(Order.status, db.func.count(Order.id))
                       .group_by(Order.status)
                       .all())
        status_counts = {row[0]: row[1] for row in status_rows}

        filters = {
            'business_type': business_type,
            'status': status,
            'start_date': start_date,
            'end_date': end_date,
            'search': search,
        }

        stats = {
            'total_orders': total_orders,
            'today_orders': today_orders,
            'pending': pending_orders,
            'total_revenue': total_revenue,
            'status_counts': status_counts,
        }

        return render_template(
            'admin_console.html',
            orders=pagination.items,
            pagination=pagination,
            stats=stats,
            filters=filters,
        )

    except Exception as e:
        current_app.logger.error(f"Error in admin_console: {e}")
        current_app.logger.error(traceback.format_exc())
        flash('An error occurred while fetching orders. See console logs for details.', 'danger')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/orders/export')
@login_required
@admin_required
def admin_orders_export():
    """
    Export filtered orders (same filters as admin_console) as a CSV.
    Optimized for Excel / Google Sheets.
    """
    import csv
    import traceback
    from io import StringIO
    from datetime import datetime, timedelta, date
    from sqlalchemy import or_
    import os
    from flask import make_response

    try:
        # --- Read filters from query string (same as admin_console) ---
        business_type  = request.args.get('business_type', 'all')
        status         = request.args.get('status', 'all')
        start_date_str = request.args.get('start_date') or ''
        end_date_str   = request.args.get('end_date') or ''
        search         = request.args.get('search', '').strip()

        query = Order.query

        # ---- Business type filter (B2B/B2C based on GSTIN) ----
        if business_type == 'b2b':
            # B2B = GSTIN present
            query = query.filter(
                Order.customer_gstin.isnot(None),
                Order.customer_gstin != ''
            )
        elif business_type == 'b2c':
            # B2C = GSTIN empty or NULL
            query = query.filter(
                or_(Order.customer_gstin.is_(None), Order.customer_gstin == '')
            )

        # ---- Status filter ----
        if status and status != 'all':
            query = query.filter(Order.status == status)

        # ---- Date range filter (end_date inclusive) ----
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                query = query.filter(Order.order_date >= start_date)
            except ValueError:
                pass

        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(Order.order_date < end_date)
            except ValueError:
                pass

        # ---- Search filter (order id, name, email, phone, pincode) ----
        if search:
            like = f"%{search}%"
            query = query.outerjoin(User, User.id == Order.user_id)
            query = query.filter(
                or_(
                    Order.id.ilike(like),
                    Order.customer_name.ilike(like),
                    Order.customer_email.ilike(like),
                    Order.shipping_name.ilike(like),
                    Order.shipping_pincode.ilike(like),
                    User.full_name.ilike(like),
                    User.email.ilike(like),
                )
            )

        orders = query.order_by(Order.order_date.desc()).all()

        # ---- Build CSV in memory ----
        output = StringIO()
        writer = csv.writer(output)

        # Header row – simple for Excel / Google Sheets
        writer.writerow([
            'Order ID',
            'Order Date',
            'Status',
            'Business Type',
            'Customer Name',
            'Customer Email',
            'Customer Phone',
            'Shipping City',
            'Shipping State',
            'Shipping Pincode',
            'Total Amount',
            'Payment Status',
            'Courier',
            'Tracking Number',
            'Cancellation Reason',
            'Items (name x qty, options, price)'
        ])

        for order in orders:
            # Shipping object (if method exists)
            shipping = order.get_shipping_address() if hasattr(order, 'get_shipping_address') else None

            business_label = 'B2B' if (getattr(order, "customer_gstin", "") or "").strip() else 'B2C'

            # Customer fields – all via getattr to avoid AttributeError
            customer_name  = getattr(order, 'customer_name', None) or getattr(order, 'shipping_name', None) or ''
            customer_email = getattr(order, 'customer_email', None) or getattr(order, 'user_email', None) or ''

            customer_phone = getattr(order, 'customer_phone', None)
            shipping_phone = getattr(shipping, 'phone', None) if shipping else None
            final_phone    = shipping_phone or customer_phone or ''

            shipping_city    = getattr(shipping, 'city', '')    if shipping else ''
            shipping_state   = getattr(shipping, 'state', '')   if shipping else ''
            shipping_pincode = getattr(shipping, 'pincode', '') if shipping else ''

            # Build a compact items string – fully defensive
            item_strings = []
            for item in getattr(order, 'items', []):
                # try several possible name fields, but NEVER crash
                item_name = (
                    getattr(item, 'name', None)
                    or getattr(item, 'artwork_name', None)
                    or getattr(item, 'title', None)
                    or getattr(item, 'product_name', None)
                    or getattr(item, 'sku', '')  # last fallback
                    or ''
                )

                quantity = getattr(item, 'quantity', None)
                size     = getattr(item, 'size', None)
                frame    = getattr(item, 'frame', None)
                glass    = getattr(item, 'glass', None)
                price    = getattr(item, 'price', None)

                parts = [item_name]

                if quantity is not None:
                    parts.append(f"x{quantity}")

                opts = []
                if size:
                    opts.append(f"Size:{size}")
                if frame:
                    opts.append(f"Frame:{frame}")
                if glass:
                    opts.append(f"Glass:{glass}")
                if opts:
                    parts.append("(" + ", ".join(opts) + ")")

                if price is not None:
                    try:
                        parts.append(f"₹{float(price):.2f}")
                    except (TypeError, ValueError):
                        pass

                item_strings.append(" ".join(parts))

            items_joined = " | ".join(item_strings)

            writer.writerow([
                order.id,
                order.order_date.strftime('%Y-%m-%d %H:%M:%S') if getattr(order, 'order_date', None) else '',
                getattr(order, 'status', '') or '',
                business_label,
                customer_name,
                customer_email,
                final_phone,
                shipping_city,
                shipping_state,
                shipping_pincode,
                f"{float(getattr(order, 'total_amount', 0) or 0):.2f}",
                getattr(order, 'payment_status', '') or '',
                getattr(order, 'courier', '') or '',
                getattr(order, 'tracking_number', '') or '',
                getattr(order, 'cancellation_reason', '') or '',
                items_joined,
            ])

        # ---- Daily auto-backup of CSV into /backups/orders ----
        try:
            backup_dir = os.path.join(current_app.root_path, 'backups', 'orders')
            os.makedirs(backup_dir, exist_ok=True)
            today_str = date.today().strftime('%Y%m%d')
            backup_path = os.path.join(backup_dir, f'orders_{today_str}.csv')
            with open(backup_path, 'w', encoding='utf-8', newline='') as f:
                f.write(output.getvalue())
        except Exception as backup_err:
            current_app.logger.error(f"CSV backup failed: {backup_err}")

        # ---- Return CSV response ----
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = 'attachment; filename=orders_export.csv'
        return response

    except Exception as e:
        current_app.logger.error(f"Error exporting orders CSV: {e}")
        current_app.logger.error(traceback.format_exc())
        flash('Failed to export orders CSV. Please try again.', 'danger')
        return redirect(url_for('admin_console'))

from collections import defaultdict
from datetime import datetime
from flask import jsonify
# (These imports are probably already at top; if yes, don't repeat)
# from decimal import Decimal
# from .models import Order, OrderItem, Address, User  # adjust if you keep models elsewhere


# =========================
# ADMIN ANALYTICS PAGE + API
# =========================

@app.route("/admin/analytics")
@login_required
@admin_required
def admin_analytics():
    return render_template("admin_analytics.html")


@app.route("/admin/analytics-data")
@login_required
@admin_required
def admin_analytics_data():
    from collections import defaultdict
    from datetime import datetime, timedelta

    # 1) Get all orders (if this crashes we *want* to see full traceback)
    orders = Order.query.order_by(Order.order_date.asc()).all()

    total_orders = len(orders)
    total_revenue = 0.0

    status_counts = {}
    by_month = defaultdict(lambda: {"orders": 0, "revenue": 0.0})
    hourly = [0] * 24
    now_utc = datetime.utcnow()
    high_risk = 0

    email_counts = {}
    pay_success = pay_failed = pay_pending = 0
    revenue_by_state = defaultdict(float)
    sku_units = defaultdict(int)
    cancelled_count = 0

    for o in orders:
        try:
            # ----- basic safe fields -----
            amount = float(getattr(o, "total_amount", 0) or 0.0)
            total_revenue += amount

            status = (getattr(o, "status", "") or "Pending Payment").strip()
            status_counts[status] = status_counts.get(status, 0) + 1

            # ----- order date handling -----
            od = getattr(o, "order_date", None)
            od_dt = None
            if isinstance(od, datetime):
                od_dt = od
            elif od:
                # try to parse string dates defensively
                try:
                    od_dt = datetime.fromisoformat(str(od))
                except Exception:
                    od_dt = None

            if od_dt is not None:
                # month bucket
                month_key = od_dt.strftime("%Y-%m")
                by_month[month_key]["orders"] += 1
                by_month[month_key]["revenue"] += amount

                # hour bucket
                h = od_dt.hour
                if 0 <= h < 24:
                    hourly[h] += 1

                # high-risk: pending/submitted more than 24h
                s_lower = status.lower()
                if ("pending payment" in s_lower or "payment submitted" in s_lower):
                    try:
                        diff_hours = (now_utc - od_dt.replace(tzinfo=None)).total_seconds() / 3600.0
                        if diff_hours > 24:
                            high_risk += 1
                    except Exception:
                        # if timezone mismatch etc – just skip high-risk for this order
                        pass

            # ----- email counts (new vs returning) -----
            email = (getattr(o, "customer_email", None)
                     or getattr(o, "user_email", None)
                     or "").strip().lower()
            if email:
                email_counts[email] = email_counts.get(email, 0) + 1

            # ----- payment status -----
            ps = (getattr(o, "payment_status", "") or "").lower()
            if "verified" in ps or "success" in ps or "paid" in ps:
                pay_success += 1
            elif "failed" in ps:
                pay_failed += 1
            else:
                pay_pending += 1

            # ----- shipping state + revenue -----
            addr = o.get_shipping_address() if hasattr(o, "get_shipping_address") else None
            state = (getattr(addr, "state", "") or "").strip()
            if state:
                revenue_by_state[state] += amount

            # ----- cancelled? -----
            if "cancelled" in status.lower():
                cancelled_count += 1

            # ----- items for SKU stats -----
            for item in getattr(o, "items", []):
                name_fallback = (
                    getattr(item, "name", None)
                    or getattr(item, "artwork_name", None)
                    or getattr(item, "title", None)
                    or getattr(item, "product_name", None)
                    or ""
                )
                sku = (getattr(item, "sku", None) or name_fallback or "").strip()
                qty = getattr(item, "quantity", 0) or 0
                if sku and qty:
                    sku_units[sku] += int(qty)

        except Exception as loop_err:
            # If one order is weird, log & skip it but DO NOT break all analytics
            app.logger.warning(f"Skipping order in analytics (id={getattr(o, 'id', '?')}): {loop_err}")

    # -------- Build final JSON structures --------
    avg_order_value = (total_revenue / total_orders) if total_orders else 0.0

    orders_by_month = [
        {"month": m, "orders": v["orders"], "revenue": round(v["revenue"], 2)}
        for m, v in sorted(by_month.items())
    ]

    hourly_orders = hourly

    new_customers = sum(1 for c in email_counts.values() if c == 1)
    returning_customers = sum(1 for c in email_counts.values() if c > 1)
    total_unique = new_customers + returning_customers
    returning_rate = (returning_customers / total_unique * 100.0) if total_unique else 0.0

    cancellation_rate = (cancelled_count / total_orders * 100.0) if total_orders else 0.0

    revenue_by_state_list = [
        {"state": st, "revenue": round(val, 2)}
        for st, val in sorted(revenue_by_state.items(), key=lambda x: x[1], reverse=True)
    ]

    top_skus = [
        {"sku": sku, "units": units}
        for sku, units in sorted(sku_units.items(), key=lambda x: x[1], reverse=True)[:10]
    ]

    predicted_repeat_orders = int(round(total_orders * (returning_rate / 100.0) * 0.6))

    best_state = revenue_by_state_list[0]["state"] if revenue_by_state_list else None
    best_hour = max(range(24), key=lambda h: hourly[h]) if any(hourly) else None

    data = {
        "summary": {
            "total_orders": total_orders,
            "total_revenue": round(total_revenue, 2),
            "avg_order_value": round(avg_order_value, 2),
        },
        "status_counts": status_counts,
        "orders_by_month": orders_by_month,
        "hourly_orders": hourly_orders,
        "new_vs_returning": {
            "new_customers": new_customers,
            "returning_customers": returning_customers,
        },
        "payment_status": {
            "success": pay_success,
            "failed": pay_failed,
            "pending": pay_pending,
        },
        "revenue_by_state": revenue_by_state_list,
        "top_skus": top_skus,
        "advanced": {
            "returning_rate": round(returning_rate, 1),
            "cancellation_rate": round(cancellation_rate, 1),
            "predicted_repeat_orders": predicted_repeat_orders,
            "high_risk_orders": high_risk,
            "best_state": best_state,
            "best_hour": best_hour,
        },
    }

    return jsonify(data)

from flask import Response, url_for

@app.route("/sitemap.xml")
def sitemap():
    pages = [
        url_for("index", _external=True),
        url_for("about", _external=True),
        url_for("contact", _external=True),
        url_for("all_products", _external=True),
    ]

    xml = ['<?xml version="1.0" encoding="UTF-8"?>']
    xml.append('<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">')

    for page in pages:
        xml.append(f"""
  <url>
    <loc>{page}</loc>
  </url>
""")

    xml.append('</urlset>')

    return Response("".join(xml), mimetype="application/xml")
@app.route("/policies")
def policies():
    return render_template("policy.html")

@app.route("/privacy-policy")
def privacy_policy():
    return render_template("privacy_policy.html")

@app.route("/terms-and-conditions")
def terms_and_conditions():
    return render_template("terms_conditions.html")
@app.route("/refund-policy")
def refund_policy():
    return render_template("refund_policy.html")

# TODO (Future): Send payment confirmation alert to admin & customer


# --- Run the App ---
if __name__ == '__main__':
    app.run(debug=False)

